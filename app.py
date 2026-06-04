from flask import Flask, render_template, request, session, jsonify, send_file, url_for, redirect, Response
from werkzeug.utils import secure_filename
import os
import time
import uuid
import json as json_module
import re
import numpy as np
from config import Config

# 导入模块
from modules.data_processor import DataProcessor
from modules.ai_codegen import AICodeGenerator
from modules.duckdb_engine import DuckDBEngine
from modules.integrity_checker import IntegrityChecker
from modules.mapping_history import save_mapping, find_match
from modules.audit_trail import log_generate, log_execute
from modules.preset_rules import get_rules, get_packs, get_custom_rules, get_rule_by_id, apply_rule, save_rule, delete_rule
from modules.sampling import get_methods, generate_sql
from modules.dify_client import DifyClient

# 用于将 session 中的映射（{标准字段名: 源字段名}）转换为前端渲染格式（{stdFieldId: 源字段名}）
JOURNAL_NAME_TO_ID = {f['name']: f['id'] for f in [
    {"id": "date", "name": "日期"},
    {"id": "summary", "name": "摘要"},
    {"id": "account_code", "name": "科目编号"},
    {"id": "subject", "name": "科目名称"},
    {"id": "company", "name": "公司名"},
    {"id": "debit", "name": "借方"},
    {"id": "credit", "name": "贷方"},
    {"id": "amount", "name": "金额"},
    {"id": "voucher_no", "name": "凭证号"},
    {"id": "department", "name": "部门"},
    {"id": "person", "name": "制单人"},
    {"id": "direction", "name": "方向"},
]}
BALANCE_NAME_TO_ID = {f['name']: f['id'] for f in [
    {"id": "company", "name": "公司名"},
    {"id": "account_code", "name": "科目编号"},
    {"id": "account_name", "name": "科目名称"},
    {"id": "beginning_debit", "name": "期初借方"},
    {"id": "beginning_credit", "name": "期初贷方"},
    {"id": "beginning", "name": "期初余额"},
    {"id": "ending_debit", "name": "期末借方"},
    {"id": "ending_credit", "name": "期末贷方"},
    {"id": "ending", "name": "期末余额"},
    {"id": "beginning_direction", "name": "期初余额方向"},
    {"id": "ending_direction", "name": "期末余额方向"},
]}

app = Flask(__name__)
app.config.from_object(Config)

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 初始化Flask-Session（使用服务端文件存储，避免客户端Cookie大小限制）
from flask_session import Session
Session(app)

app.secret_key = Config.SECRET_KEY


@app.context_processor
def inject_globals():
    """注入全局模板变量"""
    import os as _os
    commit = _os.environ.get('COMMIT_HASH', 'v1.2.0')
    return {'app_version': 'v1.2.0', 'commit_hash': commit[:8] if len(commit) > 8 else commit}





def _json_safe(obj):
    """递归转换numpy类型为原生Python类型，确保JSON序列化安全"""
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return [_json_safe(v) for v in obj]
    elif isinstance(obj, (np.integer,)):
        return int(obj)
    elif isinstance(obj, (np.floating,)):
        return float(obj)
    elif isinstance(obj, np.bool_):
        return bool(obj)
    elif isinstance(obj, np.ndarray):
        return _json_safe(obj.tolist())
    return obj

def _df_to_list(df):
    """Convert pandas DataFrame to list of dicts (may be None/empty)."""
    if df is None or df.empty:
        return []
    return [dict(zip(df.columns, row)) for row in df.itertuples(index=False)]


def _write_dicts_to_sheet(ws, rows, start_row=1, header_style=True):
    """Write a list of dicts to an openpyxl worksheet starting at start_row.
    Returns the next available row number.
    """
    if not rows:
        return start_row
    from openpyxl.styles import Font, Alignment, PatternFill
    headers = list(rows[0].keys())
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        if header_style:
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for ri, row in enumerate(rows, start_row + 1):
        for ci, h in enumerate(headers, 1):
            val = row.get(h)
            if val is None:
                val = ''
            ws.cell(row=ri, column=ci, value=val)
    return start_row + 1 + len(rows)


def _sum_col(rows, key):
    """Sum a numeric column across a list of dicts."""
    return sum(float(r.get(key, 0) or 0) for r in rows)





def get_duckdb_engine():
    """获取或创建当前会话的 DuckDB 引擎"""
    if 'session_id' not in session:
        session['session_id'] = str(uuid.uuid4())
    db_dir = app.config['DUCKDB_DIR']
    os.makedirs(db_dir, exist_ok=True)
    db_path = os.path.join(db_dir, f"{session['session_id']}.db")
    if 'duckdb_engines' not in app.extensions:
        app.extensions['duckdb_engines'] = {}
    sid = session['session_id']
    if sid not in app.extensions['duckdb_engines']:
        app.extensions['duckdb_engines'][sid] = DuckDBEngine(db_path)
    return app.extensions['duckdb_engines'][sid]


def cleanup_session_db():
    """清理当前会话的 DuckDB 数据（引擎 + 数据库文件 + 会话标记）"""
    sid = session.get('session_id')
    if not sid:
        app.logger.warning("[CLEANUP] 无 session_id，跳过清理")
        return

    # 从缓存中移除并关闭/删除引擎
    engines = app.extensions.get('duckdb_engines', {})
    engine = engines.pop(sid, None)
    if engine:
        engine.cleanup()

    # 确保数据库文件被删除（即使引擎清理失败）
    db_path = os.path.join(app.config['DUCKDB_DIR'], f"{sid}.db")
    if os.path.exists(db_path):
        try:
            os.remove(db_path)
            app.logger.info(f"[CLEANUP] 已删除数据库文件: {db_path}")
        except Exception as e:
            app.logger.error(f"[CLEANUP] 数据库文件删除失败: {e}")

    # 重置会话中的 DuckDB 相关标记
    session.pop('duckdb_imported', None)
    app.logger.info(f"[CLEANUP] 会话 {sid} DuckDB 数据已清理")


@app.route('/')
def index():
    """主页面"""
    return render_template('index.html')

@app.route('/debug/setup-demo')
def debug_setup_demo():
    """临时：为截图预先配置会话数据"""
    from modules.data_processor import DataProcessor

    je_path = "/Users/evelynn/Desktop/EY AI/成都je.xlsx"
    tb_path = "/Users/evelynn/Desktop/EY AI/成都tb.xlsx"

    if os.path.exists(je_path):
        p = DataProcessor(je_path)
        session['filepath'] = je_path
        session['data_info'] = p.process()

    if os.path.exists(tb_path):
        p = DataProcessor(tb_path)
        session['balance_filepath'] = tb_path
        session['balance_data_info'] = p.process()

    session['field_mapping'] = {
        "日期": "Effective Date",
        "摘要": "JE Description",
        "科目编号": "GL Account Number",
        "科目名称": "GL Account Name",
        "金额": "Functional Amount",
        "凭证号": "JE NUMBER",
        "公司名": "Business Unit Name",
        "借方": "Functional Debit Amount",
        "贷方": "Functional Credit Amount",
        "部门": "Business Unit",
        "制单人": "Preparer"
    }
    session['balance_field_mapping'] = {
        "公司名": "Business Unit Name",
        "科目编号": "GL Account Number",
        "科目名称": "GL Account Name",
        "期初余额": "期初",
        "期末余额": "期末"
    }

    data_info = session.get('data_info', {})
    if data_info:
        rev = {v: k for k, v in session['field_mapping'].items()}
        mapped = []
        for f in data_info.get('fields', []):
            fn = f.get('name', '')
            if fn in rev:
                mf = f.copy()
                mf['name'] = rev[fn]
                mf['original_name'] = fn
                mapped.append(mf)
            else:
                mapped.append(f.copy())
        data_info['mapped_fields'] = mapped
        preview = data_info.get('preview', [])
        if preview:
            mp = []
            for row in preview:
                mr = {}
                for k, v in row.items():
                    mr[rev.get(k, k)] = v
                mp.append(mr)
            data_info['mapped_preview'] = mp
        session['data_info'] = data_info

    balance_info = session.get('balance_data_info', {})
    if balance_info:
        rev = {v: k for k, v in session['balance_field_mapping'].items()}
        mapped = []
        for f in balance_info.get('fields', []):
            fn = f.get('name', '')
            if fn in rev:
                mf = f.copy()
                mf['name'] = rev[fn]
                mf['original_name'] = fn
                mapped.append(mf)
            else:
                mapped.append(f.copy())
        balance_info['mapped_fields'] = mapped
        preview = balance_info.get('preview', [])
        if preview:
            mp = []
            for row in preview:
                mr = {}
                for k, v in row.items():
                    mr[rev.get(k, k)] = v
                mp.append(mr)
            balance_info['mapped_preview'] = mp
        session['balance_data_info'] = balance_info

    return redirect(url_for('field_mapper_page'))

@app.route('/intro')
def intro_page():
    """产品介绍页面"""
    return render_template('intro.html')

@app.route('/upload', methods=['GET'])
def upload_page():
    """文件上传页面"""
    # 清理 DuckDB 数据和会话状态
    cleanup_session_db()
    session.pop('balance_data_info', None)
    session.pop('balance_filepath', None)
    session.pop('data_info', None)
    session.pop('filepath', None)
    session.pop('field_mapping', None)
    session.pop('balance_field_mapping', None)
    session.pop('integrity_results', None)
    session.pop('integrity_chat_history', None)
    session.pop('last_execution_result', None)
    session.pop('pending_filepath', None)
    session.pop('pending_filename', None)
    session.pop('upload_options', None)
    return render_template('upload.html')

@app.route('/field-mapper', methods=['GET'])
def field_mapper_page():
    """字段映射页面（包含序时账和科目余额表映射）"""
    data_info = session.get('data_info')
    if not data_info:
        return redirect(url_for('upload_page'))

    balance_data_info = session.get('balance_data_info')
    has_balance = bool(balance_data_info)

    # 检测是否有预填的历史映射（来自 apply-mapping-history）
    prefilled_journal = None
    prefilled_balance = None
    if session.pop('mapping_prefilled', None):
        mapping = session.get('field_mapping', {})
        if mapping:
            prefilled_journal = {}
            for std_name, source_name in mapping.items():
                fid = JOURNAL_NAME_TO_ID.get(std_name)
                if fid:
                    prefilled_journal[fid] = source_name
        balance_mapping = session.get('balance_field_mapping', {})
        if balance_mapping:
            prefilled_balance = {}
            for std_name, source_name in balance_mapping.items():
                fid = BALANCE_NAME_TO_ID.get(std_name)
                if fid:
                    prefilled_balance[fid] = source_name

    return render_template('field-mapper.html',
                         data_info=data_info,
                         balance_data_info=balance_data_info,
                         has_balance_data=has_balance,
                         prefilled_journal_mapping=prefilled_journal,
                         prefilled_balance_mapping=prefilled_balance,
                         balance_format=session.get('balance_format', 'calculated'))

@app.route('/balance-mapper', methods=['GET'])
def balance_mapper_page():
    """科目余额表字段映射已合并到字段映射页面"""
    return redirect(url_for('field_mapper_page'))

@app.route('/integrity-test', methods=['GET'])
def integrity_test_page():
    """完整性测试页面"""
    if 'data_info' not in session:
        return redirect(url_for('upload_page'))
    return render_template('integrity-test.html')

@app.route('/api/upload-balance', methods=['POST'])
def upload_balance_file():
    """API: 上传科目余额表（支持 sheet 选择和 header row 指定）"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '没有选择文件'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '没有选择文件'})

    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': '不支持的文件类型'})

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'balance_' + file.filename)
    file.save(filepath)

    # 解析可选参数：sheet_name 和 header_row
    sheet_name = request.form.get('sheet_name') or None
    header_row_raw = request.form.get('header_row')
    header_row = int(header_row_raw) - 1 if header_row_raw and header_row_raw.isdigit() else None

    try:
        processor = DataProcessor(filepath)
        data_info = processor.process(
            sheet_name=sheet_name if sheet_name and sheet_name != '__csv__' else None,
            header_row=header_row,
        )
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"科目余额表处理失败详情: {error_details}")
        return jsonify({'success': False, 'error': f'科目余额表处理失败: {str(e)}'})

    session['balance_filepath'] = filepath
    session['balance_data_info'] = data_info
    session['balance_upload_options'] = {'sheet_name': sheet_name, 'header_row': header_row}

    return jsonify(data_info)

@app.route('/api/configure-balance-fields', methods=['POST'])
def configure_balance_fields():
    """API: 配置科目余额表字段映射"""
    data = request.json
    field_mapping = data.get('field_mapping')

    if not field_mapping:
        return jsonify({'success': False, 'error': '字段映射不能为空'})

    session['balance_field_mapping'] = field_mapping

    return jsonify({
        'success': True,
        'message': '科目余额表字段映射已保存'
    })

@app.route('/api/integrity-test/run', methods=['POST'])
def run_integrity_tests():
    """API: 运行完整性测试（支持反结转和末级科目模式）"""
    try:
        data = request.get_json(silent=True) or {}
        reverse_cf = data.get('reverse_carry_forward', False)
        leaf_accts = data.get('leaf_accounts', False)

        engine = get_duckdb_engine()
        table_name = 'data' if engine.table_exists('data') else None
        balance_table = 'balance_data' if engine.table_exists('balance_data') else None

        if not table_name:
            return jsonify({'success': False, 'error': '序时账数据为空，请先上传并配置字段映射'})

        checker = IntegrityChecker(engine, journal_table=table_name, balance_table=balance_table)
        results = checker.run_all(
            reverse_carry_forward=reverse_cf,
            leaf_accounts=leaf_accts,
            balance_snapshot_table='balance_integrity',
        )

        session['integrity_results'] = results
        session.modified = True
        return jsonify(results)

    except Exception as e:
        import traceback
        app.logger.error(f"[INTEGRITY TEST ERROR] {str(e)}\n{traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': f'完整性测试执行失败: {str(e)}'
        })

@app.route('/api/integrity-test/results', methods=['GET'])
def get_integrity_results():
    """API: 获取上次完整性测试结果"""
    results = session.get('integrity_results')
    if not results:
        return jsonify({'success': False, 'error': '没有完整性测试结果'})
    return jsonify(results)

@app.route('/api/integrity-test/export', methods=['POST'])
def export_integrity_results():
    """API: 导出完整性测试结果为Excel文件（支持反结转和末级科目）"""
    try:
        data = request.get_json(silent=True) or {}
        reverse_cf = data.get('reverse_carry_forward', False)
        leaf_accts = data.get('leaf_accounts', False)

        # 获取 DuckDB 引擎
        engine = get_duckdb_engine()
        table_name = 'data' if engine.table_exists('data') else None
        balance_table = 'balance_data' if engine.table_exists('balance_data') else None

        if not table_name:
            return jsonify({'success': False, 'error': '序时账数据为空，请先上传并配置字段映射'})

        # 通过 IntegrityChecker 生成数据（DuckDB SQL 聚合）
        checker = IntegrityChecker(engine, journal_table=table_name, balance_table=balance_table)
        report = checker.export_report(reverse_carry_forward=reverse_cf, leaf_accounts=leaf_accts)

        import io
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        wb = openpyxl.Workbook()
        output = io.BytesIO()

        hfont = Font(bold=True, color='FFFFFF', size=11)
        hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        # ====== Sheet 1: 序时账完整性 ======
        ws1 = wb.active
        ws1.title = '序时账完整性'
        jrows = _df_to_list(report.get('journal'))
        if jrows:
            total = round(_sum_col(jrows, '汇总发生额'), 2)
            info = [{"指标": "汇总金额", "值": total}, {"指标": "凭证分组数", "值": len(jrows)}]
            for ri, row in enumerate(info, 1):
                for ci, (k, v) in enumerate(row.items(), 1):
                    cell = ws1.cell(row=ri, column=ci, value=v)
                    if ri == 1:
                        cell.font = hfont
                        cell.fill = hfill
            _write_dicts_to_sheet(ws1, jrows, start_row=4, header_style=True)
        else:
            ws1.cell(row=1, column=1, value="提示").font = hfont
            ws1.cell(row=2, column=1, value="无线程账数据或缺少必要字段")

        # ====== Sheet 2: 科目余额表完整性 ======
        ws2 = wb.create_sheet('科目余额表完整性')
        brows = _df_to_list(report.get('balance'))
        if brows:
            bs = round(_sum_col(brows, '期初余额'), 2)
            es = round(_sum_col(brows, '期末余额'), 2)
            amt = round(_sum_col(brows, '发生额'), 2)
            zero = '是' if abs(amt) < 0.01 else f'否（差额{amt}）'
            info = [{'指标': '期初余额合计', '值': bs},
                    {'指标': '期末余额合计', '值': es},
                    {'指标': '发生额合计(期末-期初)', '值': amt},
                    {'指标': '是否归零', '值': zero},
                    {'指标': '科目汇总数', '值': len(brows)}]
            if report.get('reverse_carry_forward_applied') and brows and '结转损益金额' in brows[0]:
                cft = round(_sum_col(brows, '结转损益金额'), 2)
                info.append({'指标': '结转损益调整额', '值': cft})
            for ri, row in enumerate(info, 1):
                for ci, (k, v) in enumerate(row.items(), 1):
                    cell = ws2.cell(row=ri, column=ci, value=v)
                    if ri == 1:
                        cell.font = hfont
                        cell.fill = hfill
            _write_dicts_to_sheet(ws2, brows, start_row=len(info) + 3, header_style=True)
            # 科目编号列设为文本格式
            if brows and '科目编号' in brows[0]:
                code_col = list(brows[0].keys()).index('科目编号') + 1
                for ri in range(len(info) + 4, len(info) + 4 + len(brows)):
                    cell = ws2.cell(row=ri, column=code_col)
                    cell.number_format = '@'
                    if cell.value is not None:
                        cell.value = str(cell.value)
        else:
            ws2.cell(row=1, column=1, value="提示").font = hfont
            ws2.cell(row=2, column=1, value="无科目余额表数据或缺少必要字段")

        # ====== Sheet 3: 交叉验证 ======
        ws3 = wb.create_sheet('交叉验证')
        crows = _df_to_list(report.get('cross_validation'))
        if crows:
            for row in crows:
                for k in list(row.keys()):
                    if k == '科目编号':
                        continue  # 科目编号保持文本，不转数字
                    if row[k] is None:
                        row[k] = 0
                    try:
                        row[k] = float(row[k])
                    except (ValueError, TypeError):
                        pass
            js = round(_sum_col(crows, '序时账发生额'), 2)
            bs2 = round(_sum_col(crows, '科目余额表发生额'), 2)
            ds = round(sum(abs(r.get('差异', 0) or 0) for r in crows), 2)
            mc = sum(1 for r in crows if abs(r.get('差异', 0) or 0) <= 0.01)
            mm = sum(1 for r in crows if abs(r.get('差异', 0) or 0) > 0.01)
            info = [{'指标': '序时账发生额合计', '值': js},
                    {'指标': '余额表发生额合计', '值': bs2},
                    {'指标': '差异绝对值合计', '值': ds},
                    {'指标': '汇总科目数', '值': len(crows)},
                    {'指标': '完全一致数', '值': mc},
                    {'指标': '存在差异数', '值': mm}]
            for ri, row in enumerate(info, 1):
                for ci, (k, v) in enumerate(row.items(), 1):
                    cell = ws3.cell(row=ri, column=ci, value=v)
                    if ri == 1:
                        cell.font = hfont
                        cell.fill = hfill
            pref = ['公司名', '科目编号', '科目名称', '科目余额表期初',
                    '科目余额表期末', '科目余额表发生额', '序时账发生额', '差异']
            allk = list(crows[0].keys())
            ordered = [c for c in pref if c in allk] + [c for c in allk if c not in pref]
            ordered_crows = [{k: row[k] for k in ordered} for row in crows]
            _write_dicts_to_sheet(ws3, ordered_crows, start_row=7, header_style=True)
            # 科目编号列设为文本格式
            code_col_idx = ordered.index('科目编号') + 1 if '科目编号' in ordered else None
            if code_col_idx:
                for ri in range(8, 8 + len(crows)):
                    cell = ws3.cell(row=ri, column=code_col_idx)
                    cell.number_format = '@'
                    if cell.value is not None:
                        cell.value = str(cell.value)
        else:
            ws3.cell(row=1, column=1, value="提示").font = hfont
            ws3.cell(row=2, column=1, value="交叉验证无法执行：科目余额表数据为空或缺少必要字段（需含：公司名、科目编号、科目名称、期初余额、期末余额），请确认已正确上传并映射科目余额表")

        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='完整性测试结果.xlsx'
        )

    except Exception as e:
        import traceback
        app.logger.error(f"[EXPORT ERROR] {str(e)}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': f'导出失败: {str(e)}'})
# ====== 完整性测试 AI 多轮对话 ======

INTEGRITY_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "get_session_info",
            "description": "查询当前会话的数据信息：已上传的表、每张表的字段列表、字段映射状态。在任何引导对话开始时调用此工具",
            "parameters": {"type": "object", "properties": {}, "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "check_journal",
            "description": "运行测试一：序时账完整性测试——按凭证分组检查借贷是否平衡",
            "parameters": {"type": "object", "properties": {}, "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "check_balance",
            "description": "运行测试二：科目余额表完整性测试——检查期初期末发生额是否归零",
            "parameters": {"type": "object", "properties": {}, "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "cross_validate",
            "description": "运行测试三：交叉验证——对比序时账和科目余额表的金额差异（需要已导入科目余额表）",
            "parameters": {"type": "object", "properties": {}, "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "run_all_tests",
            "description": "一键运行全部三项完整性测试，支持完整配置参数（方向调整、反结转、末级科目、剔除规则）。请在收集完用户所有配置后调用",
            "parameters": {
                "type": "object",
                "properties": {
                    "direction_adjustment": {
                        "type": "boolean",
                        "description": "是否启用借正贷负方向调整。如果数据中有'方向'字段且值为借/贷，设为true"
                    },
                    "reverse_carry_forward": {
                        "type": "boolean",
                        "description": "是否启用反结转处理（剔除结转损益凭证）。国产ERP通常需要"
                    },
                    "cf_account_code": {
                        "type": "string",
                        "description": "自定义未分配利润/本年利润科目号，如4103、410303。仅当reverse_carry_forward=true时有效，不填则默认用4103"
                    },
                    "cf_keywords": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "自定义摘要关键词，用于识别结转凭证。仅当reverse_carry_forward=true时有效，不填则默认['结转', '损益']"
                    },
                    "leaf_accounts": {
                        "type": "boolean",
                        "description": "是否仅用末级科目（从科目余额表中筛选末级科目）"
                    },
                    "exclude_empty_voucher": {
                        "type": "boolean",
                        "description": "是否剔除序时账中凭证编号为空的记录"
                    },
                    "exclude_balance_total": {
                        "type": "boolean",
                        "description": "是否剔除科目余额表中科目编号或名称包含'合计'的行"
                    }
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_cf_info",
            "description": "获取结转损益金额详情（反结转模式的调整依据）",
            "parameters": {"type": "object", "properties": {}, "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_leaf_info",
            "description": "获取末级科目筛选信息",
            "parameters": {"type": "object", "properties": {}, "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "export_report",
            "description": "导出完整性测试报告为Excel文件",
            "parameters": {
                "type": "object",
                "properties": {
                    "reverse_carry_forward": {
                        "type": "boolean",
                        "description": "是否启用反结转模式"
                    },
                    "leaf_accounts": {
                        "type": "boolean",
                        "description": "是否仅用末级科目"
                    }
                },
                "required": []
            }
        }
    },
]

INTEGRITY_SYSTEM_PROMPT = """你是"完整性测试助手"，担任用户的**审计数据完整性测试向导**。

## 你的角色
你不是简单的工具执行器，而是引导用户一步步完成完整性测试配置的**审计顾问**。你的核心任务是：
1. 通过提问了解用户的数据背景和需求
2. 根据用户的回答做出专业判断和建议
3. 收集完所有配置后，一次性执行测试
4. 用审计视角解读结果

## 引导流程（必须按顺序）

### 第一步：开场 + 查询数据状态
- 先调用 get_session_info 了解当前数据情况
- 根据数据状态（有序时账？有科目余额表？字段映射了哪些？）向用户简要汇报
- 然后开始引导提问

### 第二步：了解数据来源
- 问用户：「财务数据来源什么系统？」
- 常见选项：SAP / Oracle / 用友 / 金蝶 / QAD / 浪潮 / 鼎捷 / Microsoft Dynamics ...
- **国产ERP（用友/金蝶/浪潮/鼎捷等）→ 提醒可能需要反结转处理**
- **进口ERP（SAP/Oracle/QAD/Dynamics等）→ 通常不需要反结转**

### 第三步：询问借正贷负（方向调整）
- 问：「是否需要做借正贷负处理？即序时账金额是否需要根据借贷方向调整正负号？」
- 如果用户说需要：
  - 用 get_session_info 查方向字段是否已映射
  - 如果方向字段已映射 → 告知用户会自动处理
  - 如果方向字段未映射 → **提醒用户先回字段映射页面添加"方向"字段映射**
- 如果用户说不需要 → 跳过（金额原值入账）

### 第四步：询问反结转
- 问：「是否需要做反结转处理（剔除结转损益凭证）？」
- 如果用户说需要：
  - 问：「未分配利润/本年利润科目号是多少？」
  - 如果用户知道 → 记下科目号
  - 如果用户不知道 → 建议：金蝶/用友常见为4103，或者在摘要中搜索包含"结转"和"损益"的凭证
  - 用户可能给出自定义的科目号（如410303、41030101等）或自定义关键词，全部接受
- 如果用户说不需要 → 跳过

### 第五步：询问末级科目
- 问：「是否需要提取末级科目？即只使用科目余额表中底层科目，剔除上级汇总科目」
- 根据用户回答决策

### 第六步：询问剔除规则
- 问：「是否需要剔除特定数据？例如：」
  - 序时账：剔除凭证编号为空的记录？
  - 科目余额表：剔除包含"合计"、"小计"等汇总行？
  - 其他自定义剔除需求

### 第七步：汇总确认
- 将以上全部配置整理成清单，让用户确认
- 确认后调用 run_all_tests 一次性执行

### 第八步：解读结果
- 呈现测试结果，从审计视角分析异常原因
- 给出专业建议

## 工具使用规则
1. **get_session_info**：在每次对话开始时必须先调用，了解数据状态
2. **run_all_tests**：所有配置收集完毕后，统一调用此工具，一次性传入全部参数
3. **check_journal / check_balance / cross_validate**：用户要求单独运行某项时使用
4. **get_cf_info / get_leaf_info**：用户询问细节时使用
5. **export_report**：用户要求导出报告时使用

## 行为准则
1. 用简洁清晰的中文沟通，一次只问1-2个问题
2. 每一步给出专业建议，不只是问问题
3. 用户可能使用专业术语或会计科目号，灵活理解
4. 不要在用户确认前擅自执行测试
5. 不要编造工具返回数据之外的任何内容
6. 如果 get_session_info 显示无数据，引导用户先上传文件
7. **如果 integrity_tests_done 为 true，说明用户已运行过测试，直接做结果解读和差异分析，不要再问"是否运行测试"**

## 禁止行为
1. **禁止编造工具返回数据**：所有结果必须来自工具实际返回值
2. **禁止声称生成了文件或下载链接**：唯一文件输出通过 export_report 工具的返回
3. **禁止自行编造测试配置**：所有参数必须来自用户的回答
4. **不要在问完问题前就跑测试**：必须收集完完整配置再执行"""


def _execute_integrity_tool(tool_name: str, arguments: dict) -> dict:
    """执行完整性测试工具，返回格式化结果"""
    from modules.integrity_checker import IntegrityChecker
    engine = get_duckdb_engine()
    table_name = 'data' if engine.table_exists('data') else None
    balance_table = 'balance_data' if engine.table_exists('balance_data') else None

    if tool_name in ('check_journal', 'run_all_tests', 'cross_validate', 'get_leaf_info') and not table_name:
        return {"error": "序时账数据为空，请先上传并配置字段映射"}

    checker = IntegrityChecker(engine, journal_table=table_name or 'data',
                               balance_table=balance_table)

    # 单个测试工具需要与 run_all() 一致的视图预处理（trim + 方向调整）
    if tool_name in ('check_journal', 'check_balance', 'cross_validate'):
        try:
            checker._setup_trim_views()
            checker._setup_direction_views()
        except Exception:
            pass

    # 会话信息查询工具
    if tool_name == 'get_session_info':
        info = {'tables': []}
        if table_name:
            info['tables'].append('序时账')
            info['journal_columns'] = [{'name': c['name'], 'type': c['type']}
                                       for c in engine.get_schema(table_name)]
        if balance_table:
            info['tables'].append('科目余额表')
            info['balance_columns'] = [{'name': c['name'], 'type': c['type']}
                                       for c in engine.get_schema(balance_table)]
        info['journal_mapping'] = session.get('field_mapping', {})
        info['balance_mapping'] = session.get('balance_field_mapping', {})
        info['balance_format'] = session.get('balance_format', 'calculated')
        # 完整性测试状态
        integrity_results = session.get('integrity_results')
        if integrity_results:
            info['integrity_tests_done'] = True
            info['integrity_tests_all_passed'] = integrity_results.get('all_passed', False)
            s = integrity_results.get('summary', {})
            info['integrity_tests_summary'] = f"共{s.get('total', 0)}项, 完成{s.get('completed', 0)}, 跳过{s.get('skipped', 0)}, 错误{s.get('errors', 0)}"
        else:
            info['integrity_tests_done'] = False
        return {"result": json_module.dumps(info, ensure_ascii=False, default=str)}

    try:
        if tool_name == 'check_journal':
            r = checker.test_journal_integrity()
            d = r.get('details', {})
            if r.get('status') == 'error':
                return {"result": f"❌ 序时账完整性测试执行失败: {r.get('message', '')}"}
            passed = r.get('passed', False)
            lines = [f"{'✅' if passed else '❌'} 序时账完整性测试: {'通过' if passed else '未通过'}"]
            lines.append(f"  说明: {r.get('message', '')}")
            lines.append(f"  汇总金额: {d.get('total_amount', 0)}")
            lines.append(f"  分组数: 正数{d.get('positive_groups', 0)} / 负数{d.get('negative_groups', 0)} / 零值{d.get('zero_groups', 0)}")
            if not passed and d.get('groups_preview'):
                lines.append("  差异明细（非零分组）：")
                for g in d['groups_preview'][:5]:
                    amt = g.get('汇总发生额', 0)
                    if abs(amt or 0) > 0.01:
                        lines.append(f"    凭证{g.get('凭证号','')} 金额{amt}")
            return {"result": "\n".join(lines)}

        elif tool_name == 'check_balance':
            r = checker.test_balance_integrity()
            d = r.get('details', {})
            if r.get('status') == 'error':
                return {"result": f"❌ 科目余额表测试执行失败: {r.get('message', '')}"}
            if r.get('status') == 'skipped':
                return {"result": "⏭ 科目余额表测试已跳过（无科目余额表数据）"}
            passed = r.get('passed', False)
            lines = [f"{'✅' if passed else '❌'} 科目余额表完整性测试: {'通过' if passed else '未通过'}"]
            lines.append(f"  说明: {r.get('message', '')}")
            lines.append(f"  期初余额合计: {d.get('total_beginning', 0)}")
            lines.append(f"  期末余额合计: {d.get('total_ending', 0)}")
            lines.append(f"  发生额合计: {d.get('total_occurrence', 0)}")
            lines.append(f"  发生额归零检查: {'✅ 通过' if d.get('balance_check_passed') else '❌ 异常'}")
            return {"result": "\n".join(lines)}

        elif tool_name == 'cross_validate':
            r = checker.test_cross_validation()
            d = r.get('details', {})
            if r.get('status') == 'error':
                return {"result": f"❌ 交叉验证执行失败: {r.get('message', '')}"}
            if r.get('status') == 'skipped':
                return {"result": "⏭ 交叉验证已跳过（缺少科目余额表数据）"}
            passed = r.get('passed', False)
            lines = [f"{'✅' if passed else '❌'} 交叉验证: {'通过' if passed else '未通过'}"]
            lines.append(f"  说明: {r.get('message', '')}")
            lines.append(f"  汇总科目数: {d.get('total_accounts', 0)}")
            lines.append(f"  差异数量: {d.get('difference_count', 0)}")
            lines.append(f"  仅有序时账: {d.get('only_in_journal', 0)}")
            lines.append(f"  仅有余额表: {d.get('only_in_balance', 0)}")
            if d.get('diff_records'):
                lines.append("  差异明细（前5条）：")
                for rec in d['diff_records'][:5]:
                    lines.append(f"    公司{rec.get('公司名','')} 科目{rec.get('科目编号','')} {rec.get('科目名称','')} 序时账{rec.get('序时账发生额',0)} 余额表{rec.get('余额表发生额',0)} 差异{rec.get('差异',0)}")
            return {"result": "\n".join(lines)}

        elif tool_name == 'run_all_tests':
            direction_adj = arguments.get('direction_adjustment', False)
            reverse_cf = arguments.get('reverse_carry_forward', False)
            cf_account_code = arguments.get('cf_account_code') or None
            cf_keywords = arguments.get('cf_keywords') or None
            leaf = arguments.get('leaf_accounts', False)
            exclude_empty = arguments.get('exclude_empty_voucher', False)
            exclude_total = arguments.get('exclude_balance_total', False)

            orig_jt = checker.journal_table
            orig_bt = checker.balance_table
            conn = checker.engine._conn
            config_items = []

            # 1) TRIM
            checker._setup_trim_views()
            config_items.append("✓ 文本去空格")

            # 2) 方向调整（按参数）
            if direction_adj:
                checker._setup_direction_views()
                config_items.append("✓ 借正贷负方向调整")
            else:
                config_items.append("○ 跳过方向调整")

            # 3) 自定义剔除规则
            if exclude_empty and table_name:
                conn.execute(f'''
                    CREATE OR REPLACE TEMP VIEW _chat_ex_j AS
                    SELECT * FROM "{checker.journal_table}"
                    WHERE "凭证号" IS NOT NULL AND CAST("凭证号" AS VARCHAR) != ''
                ''')
                checker.journal_table = '_chat_ex_j'
                config_items.append("✓ 剔除凭证号为空")
            if exclude_total and balance_table:
                conn.execute(f'''
                    CREATE OR REPLACE TEMP VIEW _chat_ex_b AS
                    SELECT * FROM "{checker.balance_table}"
                    WHERE (CAST("科目编号" AS VARCHAR) NOT LIKE '%合计%')
                      AND (CAST("科目名称" AS VARCHAR) NOT LIKE '%合计%')
                ''')
                checker.balance_table = '_chat_ex_b'
                config_items.append("✓ 剔除合计行")

            # 4) 末级科目
            if leaf and balance_table:
                try:
                    checker._setup_leaf_account_views()
                    checker.balance_table = 'balance_leaf'
                    config_items.append("✓ 末级科目筛选")
                except Exception as e:
                    config_items.append(f"⚠ 末级科目筛选失败: {e}")

            # 5) 反结转（自定义科目号/关键词）
            if reverse_cf and table_name:
                try:
                    jt = checker.journal_table
                    bt = checker.balance_table
                    j_cols = {c['name'] for c in (engine.get_schema(jt) or [])}
                    cf_conds = []
                    if '科目编号' in j_cols:
                        code = cf_account_code or '4103'
                        cf_conds.append(f'CAST("科目编号" AS VARCHAR) = \'{code}\'')
                    kw_list = cf_keywords or ['结转', '损益']
                    if '摘要' in j_cols and kw_list:
                        kw_cond = ' AND '.join(f'"摘要" LIKE \'%{kw}%\''
                                                for kw in kw_list)
                        cf_conds.append(f'({kw_cond})')
                    if cf_conds:
                        w = ' OR '.join(cf_conds)
                        conn.execute(f'''
                            CREATE OR REPLACE TEMP VIEW _chat_cf_v AS
                            SELECT DISTINCT "公司名","日期","凭证号" FROM "{jt}" WHERE {w}
                        ''')
                        # cf_amounts（按方向处理金额）
                        if '方向' in j_cols:
                            conn.execute(f'''
                                CREATE OR REPLACE TEMP VIEW _chat_cf_a AS
                                SELECT d."公司名",CAST(d."科目编号" AS VARCHAR)"科目编号",d."科目名称",
                                  SUM(CASE WHEN d."方向" IN ('借','Debit','D') THEN CAST(d."金额" AS DOUBLE) ELSE -CAST(d."金额" AS DOUBLE) END)"结转损益金额"
                                FROM "{jt}" d JOIN _chat_cf_v v ON d."公司名"=v."公司名" AND d."日期"=v."日期" AND d."凭证号"=v."凭证号"
                                GROUP BY d."公司名",d."科目编号",d."科目名称"
                            ''')
                        else:
                            conn.execute(f'''
                                CREATE OR REPLACE TEMP VIEW _chat_cf_a AS
                                SELECT d."公司名",CAST(d."科目编号" AS VARCHAR)"科目编号",d."科目名称",
                                  SUM(CAST(d."金额" AS DOUBLE))"结转损益金额"
                                FROM "{jt}" d JOIN _chat_cf_v v ON d."公司名"=v."公司名" AND d."日期"=v."日期" AND d."凭证号"=v."凭证号"
                                GROUP BY d."公司名",d."科目编号",d."科目名称"
                            ''')
                        conn.execute(f'''
                            CREATE OR REPLACE TEMP VIEW _chat_j_filt AS
                            SELECT d.* FROM "{jt}" d
                            LEFT JOIN _chat_cf_v v ON d."公司名"=v."公司名" AND d."日期"=v."日期" AND d."凭证号"=v."凭证号"
                            WHERE v."公司名" IS NULL
                        ''')
                        checker.journal_table = '_chat_j_filt'
                        if bt and engine.table_exists(bt):
                            b_cols = {c['name'] for c in (engine.get_schema(bt) or [])}
                            if '期初余额' in b_cols and '期末余额' in b_cols:
                                conn.execute(f'''
                                    CREATE OR REPLACE TEMP VIEW _chat_b_adj AS
                                    SELECT b."公司名",CAST(b."科目编号" AS VARCHAR)"科目编号",b."科目名称",
                                      SUM(CAST(b."期初余额" AS DOUBLE))"期初余额",
                                      SUM(CAST(b."期末余额" AS DOUBLE))-COALESCE(SUM(cf."结转损益金额"),0)"期末余额",
                                      SUM(CAST(b."期末余额" AS DOUBLE))-COALESCE(SUM(cf."结转损益金额"),0)-SUM(CAST(b."期初余额" AS DOUBLE))"发生额"
                                    FROM "{bt}" b
                                    LEFT JOIN _chat_cf_a cf ON CAST(b."公司名" AS VARCHAR)=CAST(cf."公司名" AS VARCHAR) AND CAST(b."科目编号" AS VARCHAR)=CAST(cf."科目编号" AS VARCHAR)
                                    GROUP BY b."公司名",b."科目编号",b."科目名称"
                                ''')
                                checker.balance_table = '_chat_b_adj'
                        config_items.append(f"✓ 反结转（科目{cf_account_code or '4103'}）")
                except Exception as e:
                    config_items.append(f"⚠ 反结转失败: {e}")

            # 6) 执行三项测试
            try:
                results = {
                    'journal_test': checker.test_journal_integrity(),
                    'balance_test': checker.test_balance_integrity(),
                    'cross_test': checker.test_cross_validation(),
                }
            finally:
                checker.journal_table = orig_jt
                checker.balance_table = orig_bt
                for _v in ['_chat_ex_j','_chat_ex_b','_chat_cf_v','_chat_cf_a','_chat_j_filt','_chat_b_adj']:
                    try:
                        conn.execute(f'DROP VIEW IF EXISTS "{_v}"')
                    except Exception:
                        pass
                checker._drop_cf_views()
                checker._drop_leaf_views()
                checker._drop_direction_views()
                checker._drop_trim_views()

            total = 3
            completed = sum(1 for r in results.values() if r['status'] == 'completed')
            errors = sum(1 for r in results.values() if r['status'] == 'error')
            skipped = sum(1 for r in results.values() if r['status'] == 'skipped')
            all_passed = all(r.get('passed', False) for r in results.values() if r['status'] == 'completed')
            r = {'success': True, 'summary': {'total': total, 'completed': completed, 'errors': errors, 'skipped': skipped},
                 'results': results, 'all_passed': all_passed}
            session['integrity_results'] = r
            session.modified = True

            lines = ["=== 完整性测试结果 ==="]
            lines.append(f"总{total}项 | ✅通过{completed} | ❌未通过{total-completed-errors-skipped} | ⏭跳过{skipped} | ❌错误{errors}")
            for tk, tl in [('journal_test', '序时账'), ('balance_test', '科目余额表'), ('cross_test', '交叉验证')]:
                t = results.get(tk, {})
                st = t.get('status', '')
                if st == 'completed':
                    lines.append(f"  {tl}: {'✅' if t.get('passed') else '❌'} {t.get('message','')}")
                elif st == 'skipped':
                    lines.append(f"  {tl}: ⏭ 跳过")
                elif st == 'error':
                    lines.append(f"  {tl}: ❌ {t.get('message','')}")
            if all_passed:
                lines.append("\n🎉 全部测试通过！")
            lines.append("\n--- 本次测试配置 ---")
            lines.extend(config_items)
            return {"result": "\n".join(lines)}

        elif tool_name == 'get_cf_info':
            r = checker.get_cf_info()
            if r.get('error'):
                return {"result": f"结转损益信息获取失败: {r['error']}"}
            lines = ["=== 结转损益金额详情 ==="]
            lines.append(f"  公司: {r.get('company', '')}")
            lines.append(f"  科目编号: {r.get('account_code', '')}")
            lines.append(f"  科目名称: {r.get('account_name', '')}")
            lines.append(f"  结转金额: {r.get('total_amount', 0)}")
            lines.append(f"  匹配类型: {r.get('match_type', '')}")
            return {"result": "\n".join(lines)}

        elif tool_name == 'get_leaf_info':
            r = checker.get_leaf_info()
            if r.get('error'):
                return {"result": f"末级科目信息获取失败: {r['error']}"}
            lines = ["=== 末级科目筛选信息 ==="]
            lines.append(f"  末级科目数: {r.get('leaf_count', 0)}")
            lines.append(f"  非末级科目数: {r.get('non_leaf_count', 0)}")
            return {"result": "\n".join(lines)}

        elif tool_name == 'export_report':
            report = checker.export_report(
                reverse_carry_forward=arguments.get('reverse_carry_forward', False),
                leaf_accounts=arguments.get('leaf_accounts', False))
            import io, openpyxl, base64
            wb = openpyxl.Workbook()
            sheet_keys = ['journal', 'balance', 'cross_validation']
            sheet_names = {'journal': '序时账', 'balance': '科目余额表', 'cross_validation': '交叉验证'}
            first = True
            for key in sheet_keys:
                rows = _df_to_list(report.get(key))
                if not rows:
                    continue
                ws = wb.active if first else wb.create_sheet()
                first = False
                ws.title = sheet_names.get(key, key[:31])
                for ri, row in enumerate(rows, 1):
                    for ci, (k, v) in enumerate(row.items(), 1):
                        ws.cell(row=ri, column=ci, value=v if v is not None else '')
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            b64 = base64.b64encode(output.getvalue()).decode()
            return {"result": "报告已生成（含序时账、科目余额表、交叉验证三个Sheet），请使用下方下载按钮获取", "file": b64, "filename": "integrity_report.xlsx"}

    except Exception as e:
        return {"error": f"工具执行异常: {str(e)}"}


@app.route('/api/integrity-chat', methods=['POST'])
def integrity_chat():
    """多轮对话：完整性测试助手"""
    data = request.get_json(silent=True) or {}
    message = data.get('message', '').strip()
    if not message:
        return jsonify({'success': False, 'error': '消息不能为空'})

    api_key = session.get('api_key')
    if not api_key:
        return jsonify({'success': False, 'error': '完整性测试助手暂不可用，请直接使用上方配置运行测试'})

    try:
        provider = session.get('ai_provider', 'deepseek')
        model = session.get('ai_model')
        plain_key = api_key  # 功能暂不可用
        model_name = 'deepseek-chat'
        api_url = 'https://api.deepseek.com/v1/chat/completions'

        history = session.get('integrity_chat_history', [])
        messages = [{"role": "system", "content": INTEGRITY_SYSTEM_PROMPT}]
        messages.extend(history)
        messages.append({"role": "user", "content": message})

        import requests as http_req

        def call_ai(msgs, max_rounds=5):
            file_data = None
            file_name = None
            for _ in range(max_rounds):
                resp = http_req.post(api_url, headers={
                    "Authorization": f"Bearer {plain_key}",
                    "Content-Type": "application/json"
                }, json={
                    "model": model_name,
                    "messages": msgs,
                    "tools": INTEGRITY_TOOLS,
                    "temperature": 0.3,
                    "max_tokens": 2000,
                }, timeout=60)

                if resp.status_code != 200:
                    raise Exception(f"API请求失败: {resp.status_code}")

                result = resp.json()
                choice = result['choices'][0]
                msg = choice['message']

                if not msg.get('tool_calls'):
                    return msg.get('content', ''), file_data, file_name

                msgs.append({"role": "assistant", "content": msg.get('content') or None, "tool_calls": msg['tool_calls']})

                for tc in msg['tool_calls']:
                    fn = tc['function']
                    tool_result = _execute_integrity_tool(fn['name'], json_module.loads(fn.get('arguments', '{}')))
                    if isinstance(tool_result, dict) and 'file' in tool_result:
                        file_data = tool_result['file']
                        file_name = tool_result.get('filename', 'integrity_report.xlsx')
                    msgs.append({
                        "role": "tool",
                        "tool_call_id": tc['id'],
                        "content": json_module.dumps(tool_result, ensure_ascii=False)
                    })

            return "抱歉，对话处理超过最大轮次，请重试。", file_data, file_name

        reply, file_data, file_name = call_ai(messages)

        # 兜底：用户要求导出但 AI 没调 export_report 工具时，后端强制导出
        if not file_data:
            export_keywords = ['导出', '报告', '下载']
            if any(kw in message for kw in export_keywords):
                try:
                    tool_result = _execute_integrity_tool('export_report', {})
                    if isinstance(tool_result, dict) and 'file' in tool_result:
                        file_data = tool_result['file']
                        file_name = tool_result.get('filename', 'integrity_report.xlsx')
                except Exception:
                    pass

        # export_report 被调用（或强制导出）后，用固定消息避免 AI 幻觉
        if file_data:
            reply = "✅ 完整性测试报告已生成，请点击下方下载按钮获取。"

        trimmed_history = history + [
            {"role": "user", "content": message},
            {"role": "assistant", "content": reply}
        ]
        if len(trimmed_history) > 20:
            trimmed_history = trimmed_history[-20:]
        session['integrity_chat_history'] = trimmed_history

        result_data = {'success': True, 'response': reply}
        if file_data:
            result_data['file_data'] = file_data
            result_data['file_name'] = file_name
        return jsonify(result_data)

    except Exception as e:
        import traceback
        app.logger.error(f"[INTEGRITY-CHAT] {str(e)}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': f'对话处理失败: {str(e)}'})


@app.route('/api/integrity-test/ai-analyze', methods=['POST'])
def ai_analyze_integrity():
    """AI 分析完整性测试异常结果 — 从审计视角分析原因并给出建议"""
    results = session.get('integrity_results')

    # 如果还没有测试结果，自动重新跑一遍
    if not results:
        try:
            engine = get_duckdb_engine()
            table_name = 'data' if engine.table_exists('data') else None
            balance_table = 'balance_data' if engine.table_exists('balance_data') else None
            if table_name:
                from modules.integrity_checker import IntegrityChecker
                checker = IntegrityChecker(engine, journal_table=table_name, balance_table=balance_table)
                results = checker.run_all()
                session['integrity_results'] = results
                session.modified = True
        except Exception as e:
            app.logger.error(f"[AI-ANALYZE] auto re-run failed: {e}")

    if not results:
        return jsonify({'success': False, 'error': '没有完整性测试结果，请先运行测试'})
    if results.get('all_passed'):
        return jsonify({'success': False, 'error': '所有测试均已通过'})

    try:
        # ---- 构建 prompt ----
        lines = [
            "你是一名资深的财务审计专家。以下是财务数据完整性测试的结果，请从审计专业角度逐项分析异常原因，并给出后续建议。",
            "",
            "## 测试汇总",
        ]
        summary = results.get('summary', {})
        lines.append(f"- 总测试数: {summary.get('total', 3)}")
        lines.append(f"- 完成: {summary.get('completed', 0)}  错误: {summary.get('errors', 0)}  跳过: {summary.get('skipped', 0)}")
        if results.get('reverse_carry_forward_applied'):
            lines.append("- 已启用反结转模式")
        if results.get('leaf_accounts_applied'):
            lines.append("- 已启用末级科目筛选")

        cf_info = results.get('carry_forward_info')
        if cf_info and cf_info.get('error'):
            lines.append(f"\n⚠ 反结转执行异常: {cf_info['error']}")
        lines.append("")

        for test_key, test_label in [
            ('journal_test', '测试一：序时账完整性'),
            ('balance_test', '测试二：科目余额表完整性'),
            ('cross_test', '测试三：交叉验证'),
        ]:
            test = results.get('results', {}).get(test_key, {})
            if not test or test.get('status') != 'completed' or test.get('passed'):
                continue

            details = test.get('details', {})
            lines.append(f"### {test_label}")
            lines.append(f"说明: {test.get('message', '')}")

            if test_key == 'journal_test':
                lines.append(f"汇总金额: {details.get('total_amount', 0)}")
                lines.append(f"正数/负数/零值分组: {details.get('positive_groups', 0)} / {details.get('negative_groups', 0)} / {details.get('zero_groups', 0)}")
                preview = details.get('groups_preview', [])
                non_zero = [g for g in preview if abs(g.get('汇总发生额', 0) or 0) > 0.01]
                if non_zero:
                    lines.append(f"非零分组（前 {min(10, len(non_zero))} 条）：")
                    for g in non_zero[:10]:
                        lines.append(f"  公司:{g.get('公司名','')} 凭证:{g.get('凭证号','')} 日期:{g.get('日期','')} 金额:{g.get('汇总发生额',0)}")
                    total_groups = details.get('total_groups', 0)
                    if total_groups > 10:
                        lines.append(f"  ...共 {total_groups} 个分组")

            elif test_key == 'balance_test':
                lines.append(f"期初余额合计: {details.get('total_beginning', 0)}")
                lines.append(f"期末余额合计: {details.get('total_ending', 0)}")
                lines.append(f"发生额合计(期末-期初): {details.get('total_occurrence', 0)}")
                lines.append(f"发生额归零检查: {'通过' if details.get('balance_check_passed') else '异常'}")

            elif test_key == 'cross_test':
                lines.append(f"汇总科目数: {details.get('total_accounts', 0)}")
                lines.append(f"差异数量: {details.get('difference_count', 0)}")
                lines.append(f"仅有序时账: {details.get('only_in_journal', 0)}")
                lines.append(f"仅有余额表: {details.get('only_in_balance', 0)}")
                lines.append(f"序时账发生额合计: {details.get('journal_total_amount', 0)}")
                lines.append(f"余额表发生额合计: {details.get('balance_total_occurrence', 0)}")
                diff_records = details.get('diff_records', [])
                if diff_records:
                    lines.append(f"差异明细（前 {min(10, len(diff_records))} 条，共 {details.get('difference_count', 0)} 条）：")
                    for r in diff_records[:10]:
                        lines.append(f"  公司:{r.get('公司名','')} 科目:{r.get('科目编号','')} {r.get('科目名称','')}  序时账发生额:{r.get('序时账发生额',0)}  余额表发生额:{r.get('余额表发生额',0)}  差异:{r.get('差异',0)}")
            lines.append("")

        lines.extend([
            "请按以下格式输出分析结果，每个有异常的测试单独分析：",
            "",
            "### [测试名称]",
            "**异常情况**：（简要描述问题）",
            "**可能原因**：（从审计专业视角分析，例如）",
            "- 财务系统导出问题（导出范围不完整、借贷方向不一致、未包含所有期间等）",
            "- 数据自身问题（某月未做损益结转、凭证借贷不平时手工调账、缺少部分科目等）",
            "- 用户配置问题（字段映射错误、期初期末余额方向未正确处理、反结转条件不适配等）",
            "- 会计处理差异（SAP/金蝶/用友等不同系统的特有处理方式）",
            "**建议后续操作**：（列出 1-3 条可操作的审计建议）",
            "",
            "注意：已通过或跳过的测试不要分析。语言专业简洁，用中文。",
        ])
        prompt = "\n".join(lines)

        # ---- 通过 Dify 代理调用 AI ----
        dify = _get_dify_client()
        content = dify.chat(
            "你是一名资深的财务审计专家，擅长从审计视角分析财务数据问题。",
            prompt,
            timeout=60,
        )
        return jsonify({'success': True, 'analysis': content.strip()})

    except Exception as e:
        import traceback
        app.logger.error(f"[AI-ANALYZE-INTEGRITY] {str(e)}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': f'AI 分析失败: {str(e)}'})


@app.route('/api/configure-fields', methods=['POST'])
def configure_fields():
    """API: 配置字段映射（序时账 + 科目余额表）"""
    data = request.json
    field_mapping = data.get('field_mapping')
    balance_field_mapping = data.get('balance_field_mapping')

    if not field_mapping:
        return jsonify({'success': False, 'error': '序时账字段映射不能为空'})

    # 存储序时账映射
    session['field_mapping'] = field_mapping
    app.logger.info(f"[CONFIGURE_FIELDS] field_mapping received: {field_mapping}")

    # 存储手动填写的字段（如手动输入的公司名）
    manual_fills = data.get('manual_fills', {})
    balance_manual_fills = data.get('balance_manual_fills', {})
    session['manual_fills'] = manual_fills
    session['balance_manual_fills'] = balance_manual_fills
    if manual_fills:
        app.logger.info(f"[CONFIGURE_FIELDS] manual_fills: {manual_fills}")
    if balance_manual_fills:
        app.logger.info(f"[CONFIGURE_FIELDS] balance_manual_fills: {balance_manual_fills}")

    # 更新 data_info 中的 mapped_fields 和 mapped_preview
    data_info = session.get('data_info', {})
    app.logger.info(f"[CONFIGURE_FIELDS] data_info keys before: {list(data_info.keys()) if data_info else 'EMPTY'}")
    if data_info and field_mapping:
        reverse_mapping = {v: k for k, v in field_mapping.items()}
        app.logger.info(f"[CONFIGURE_FIELDS] reverse_mapping: {reverse_mapping}")

        # 映射字段信息
        mapped_fields = []
        for field in data_info.get('fields', []):
            field_name = field.get('name', '')
            if field_name in reverse_mapping:
                mapped_field = field.copy()
                mapped_field['name'] = reverse_mapping[field_name]
                mapped_field['original_name'] = field_name
                mapped_fields.append(mapped_field)
            else:
                mapped_fields.append(field.copy())
        data_info['mapped_fields'] = mapped_fields
        app.logger.info(f"[CONFIGURE_FIELDS] mapped_fields: {[f.get('name') for f in mapped_fields]}")

        # 映射预览数据
        preview = data_info.get('preview', [])
        if preview:
            mapped_preview = []
            for row in preview:
                mapped_row = {}
                for key, value in row.items():
                    mapped_key = reverse_mapping.get(key, key)
                    mapped_row[mapped_key] = value
                mapped_preview.append(mapped_row)
            data_info['mapped_preview'] = mapped_preview
            app.logger.info(f"[CONFIGURE_FIELDS] mapped_preview[0] keys: {list(mapped_preview[0].keys()) if mapped_preview else 'EMPTY'}")

        session['data_info'] = data_info
        app.logger.info(f"[CONFIGURE_FIELDS] session data_info keys after: {list(session.get('data_info', {}).keys())}")

    # 存储科目余额表格式
    balance_format = data.get('balance_format', 'calculated')
    session['balance_format'] = balance_format
    app.logger.info(f"[CONFIGURE_FIELDS] balance_format: {balance_format}")

    # 存储科目余额表映射（如果有），同时更新 balance_data_info
    if balance_field_mapping:
        session['balance_field_mapping'] = balance_field_mapping

        balance_info = session.get('balance_data_info', {})
        if balance_info:
            rev_balance = {v: k for k, v in balance_field_mapping.items()}
            mapped_fields = []
            for field in balance_info.get('fields', []):
                fname = field.get('name', '')
                if fname in rev_balance:
                    mf = field.copy()
                    mf['name'] = rev_balance[fname]
                    mf['original_name'] = fname
                    mapped_fields.append(mf)
                else:
                    mapped_fields.append(field.copy())
            balance_info['mapped_fields'] = mapped_fields

            bpreview = balance_info.get('preview', [])
            if bpreview:
                mapped_preview = []
                for row in bpreview:
                    mrow = {}
                    for key, value in row.items():
                        mrow[rev_balance.get(key, key)] = value
                    mapped_preview.append(mrow)
            balance_info['mapped_preview'] = mapped_preview

        session['balance_data_info'] = balance_info

    has_balance = bool(session.get('balance_data_info'))

    # ---- 导入数据到 DuckDB ----
    # 重新导入数据时清除旧对话历史
    session.pop('integrity_chat_history', None)
    import_success = False
    try:
        engine = get_duckdb_engine()
        filepath = session.get('filepath')
        field_mapping = session.get('field_mapping')
        upload_opts = session.get('upload_options', {})
        sheet_name = upload_opts.get('sheet_name')
        header_row = upload_opts.get('header_row')
        constant_cols = session.get('manual_fills') or None
        if filepath and field_mapping:
            ext = os.path.splitext(filepath)[1].lower()
            reverse_mapping = {v: k for k, v in field_mapping.items()}
            if ext == '.csv':
                rows = engine.import_csv(filepath, 'data', reverse_mapping, header_row=header_row,
                                         constant_columns=constant_cols)
            elif ext == '.xlsx':
                rows = engine.import_xlsx(filepath, 'data', reverse_mapping,
                                          sheet_name=sheet_name, header_row=header_row,
                                          constant_columns=constant_cols)
            else:
                rows = 0
            app.logger.info(f"[DUCKDB] 序时账已导入: {filepath} → {rows} 行")
            if rows == 0:
                app.logger.warning(f"[DUCKDB] 序时账导入后行数为 0，请检查源文件")
            import_success = True

            # 如果序时账有借方和贷方但没有金额字段，自动计算金额 = 借方 - 贷方
            try:
                j_cols = {c['name'] for c in engine.get_schema('data')}
                has_debit = '借方' in j_cols
                has_credit = '贷方' in j_cols
                has_amount = '金额' in j_cols
                if has_debit and has_credit and not has_amount:
                    engine._conn.execute('''
                        ALTER TABLE "data" ADD COLUMN "金额" DOUBLE
                    ''')
                    engine._conn.execute('''
                        UPDATE "data" SET "金额" = COALESCE(TRY_CAST("借方" AS DOUBLE),0) - COALESCE(TRY_CAST("贷方" AS DOUBLE),0)
                    ''')
                    app.logger.info("[DUCKDB] 已从借贷方自动计算序时账金额")
            except Exception as e:
                app.logger.warning(f"[DUCKDB] 序时账金额计算失败: {e}")
        else:
            app.logger.warning(f"[DUCKDB] 跳过序时账导入: filepath={filepath}, field_mapping={'有' if field_mapping else '无'}")

        balance_filepath = session.get('balance_filepath')
        balance_field_mapping = session.get('balance_field_mapping')
        balance_constant_cols = session.get('balance_manual_fills') or None
        balance_upload_opts = session.get('balance_upload_options', {})
        balance_sheet_name = balance_upload_opts.get('sheet_name')
        balance_header_row = balance_upload_opts.get('header_row')
        if balance_filepath and balance_field_mapping:
            ext = os.path.splitext(balance_filepath)[1].lower()
            rev_balance = {v: k for k, v in balance_field_mapping.items()}
            if ext == '.csv':
                engine.import_csv(balance_filepath, 'balance_data', rev_balance,
                                  header_row=balance_header_row,
                                  constant_columns=balance_constant_cols)
            elif ext == '.xlsx':
                engine.import_xlsx(balance_filepath, 'balance_data', rev_balance,
                                   sheet_name=balance_sheet_name, header_row=balance_header_row,
                                   constant_columns=balance_constant_cols)
            app.logger.info(f"[DUCKDB] 科目余额表已导入: {balance_filepath}")

            # 如果余额表格式为借贷方，自动计算期初余额和期末余额
            if session.get('balance_format') == 'debit_credit':
                try:
                    b_cols = {c['name'] for c in engine.get_schema('balance_data')}
                    has_beg_dr = '期初借方' in b_cols
                    has_beg_cr = '期初贷方' in b_cols
                    has_end_dr = '期末借方' in b_cols
                    has_end_cr = '期末贷方' in b_cols
                    if has_beg_dr and has_beg_cr:
                        engine._conn.execute('''
                            ALTER TABLE "balance_data" ADD COLUMN "期初余额" DOUBLE
                        ''')
                        engine._conn.execute('''
                            UPDATE "balance_data" SET "期初余额" = COALESCE(TRY_CAST("期初借方" AS DOUBLE),0) - COALESCE(TRY_CAST("期初贷方" AS DOUBLE),0)
                        ''')
                        app.logger.info("[DUCKDB] 已从借贷方计算期初余额")
                    if has_end_dr and has_end_cr:
                        engine._conn.execute('''
                            ALTER TABLE "balance_data" ADD COLUMN "期末余额" DOUBLE
                        ''')
                        engine._conn.execute('''
                            UPDATE "balance_data" SET "期末余额" = COALESCE(TRY_CAST("期末借方" AS DOUBLE),0) - COALESCE(TRY_CAST("期末贷方" AS DOUBLE),0)
                        ''')
                        app.logger.info("[DUCKDB] 已从借贷方计算期末余额")
                except Exception as e:
                    app.logger.warning(f"[DUCKDB] 借贷方计算余额失败: {e}")
    except Exception as e:
        app.logger.error(f"[DUCKDB IMPORT ERROR] {str(e)}")
        # DuckDB 导入失败时清理残留的 .db 文件
        cleanup_session_db()
        return jsonify({
            'success': False,
            'error': f'数据导入DuckDB失败: {str(e)}，字段映射已保存，请检查源文件后重试'
        })

    session['duckdb_imported'] = import_success

    if not import_success:
        app.logger.warning("[DUCKDB] 序时账未导入（无文件或无字段映射），字段映射配置已保存")

    # 映射成功后自动保存历史记录
    if import_success:
        try:
            filepath = session.get('filepath')
            if filepath and os.path.exists(filepath):
                data_info = session.get('data_info', {})
                cols = [f.get('name', '') for f in data_info.get('fields', [])]
                balance_info = session.get('balance_data_info', {})
                balance_cols = [f.get('name', '') for f in balance_info.get('fields', [])]
                save_mapping(
                    filename=os.path.basename(filepath),
                    file_size=os.path.getsize(filepath),
                    field_mapping=field_mapping,
                    balance_field_mapping=balance_field_mapping or {},
                    original_columns=cols,
                    balance_original_columns=balance_cols,
                )
                app.logger.info(f"[HISTORY] 映射历史已保存: {os.path.basename(filepath)}")
        except Exception as e:
            app.logger.error(f"[HISTORY] 保存映射历史失败: {e}")

    return jsonify({
        'success': True,
        'message': '字段映射已保存并应用',
        'has_balance_data': has_balance
    })


@app.route('/api/auto-map-fields', methods=['POST'])
def auto_map_fields():
    """AI 自动字段映射 — 调用 AI 根据列名、类型、样本值推荐映射"""
    data = request.json

    journal_fields = data.get('journal_fields', [])
    journal_standard = data.get('journal_standard_fields', [])
    balance_fields = data.get('balance_fields', [])
    balance_standard = data.get('balance_standard_fields', [])

    if (not journal_fields or not journal_standard) and (not balance_fields or not balance_standard):
        return jsonify({'success': False, 'error': '字段信息不完整'})

    try:
        # Build prompt
        lines = [
            "你是一个数据工程师，负责将用户上传的Excel列映射到标准字段。",
            "根据列名、数据类型和样本值推断每列的含义，映射到最合适的标准字段。",
        ]

        has_j = bool(journal_fields and journal_standard)
        has_b = bool(balance_fields and balance_standard)

        if has_j:
            lines.extend([
                "",
                "## 用户序时账列",
            ])
            for f in journal_fields:
                lines.append(f"- {f['name']} (类型: {f['type']}, 样本: {f.get('sample', 'N/A')})")
            lines.extend(["", "## 标准字段（序时账）"])
            for f in journal_standard:
                lines.append(f"- {f['id']}: {f['name']} — {f['description']} (类型: {f['type']})")

        if has_b:
            lines.extend(["", "## 用户科目余额表列"])
            for f in balance_fields:
                lines.append(f"- {f['name']} (类型: {f['type']}, 样本: {f.get('sample', 'N/A')})")
            lines.extend(["", "## 标准字段（科目余额表）"])
            for f in balance_standard:
                lines.append(f"- {f['id']}: {f['name']} — {f['description']} (类型: {f['type']})")

        return_format = []
        if has_j:
            return_format.append('"journal_mapping": { "date": "用户列名", ... }')
        if has_b:
            return_format.append(f'"balance_mapping": {{ "beginning": "用户列名", "ending": "用户列名", ... }}')
        lines.extend([
            "",
            f'返回纯 JSON，不要 Markdown 包裹。格式：{{ {", ".join(return_format)} }}',
            "映射不确认的字段就省略。",
        ])
        prompt = "\n".join(lines)

        # 通过 Dify 代理调用 AI
        dify = _get_dify_client()
        content = dify.chat(
            "你只返回 JSON，不加 Markdown 代码块。",
            prompt,
            timeout=30,
        )

        # Strip markdown code block wrappers if present
        if content.startswith('```'):
            content = re.sub(r'^```(?:json)?\s*', '', content)
            content = re.sub(r'\s*```$', '', content)
            content = content.strip()

        # Find the first JSON object
        brace_start = content.find('{')
        brace_end = content.rfind('}')
        if brace_start >= 0 and brace_end >= 0:
            content = content[brace_start:brace_end + 1]

        mapping_result = json_module.loads(content)

        return jsonify({
            'success': True,
            'journal_mapping': mapping_result.get('journal_mapping', {}),
            'balance_mapping': mapping_result.get('balance_mapping', {})
        })

    except Exception as e:
        import traceback
        app.logger.error(f"[AUTO-MAP-FIELDS] {str(e)}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': f'AI 自动映射失败: {str(e)}'})


@app.route('/api/mapping-history/check', methods=['GET'])
def check_mapping_history():
    """API: 检查当前上传文件是否有历史映射记录"""
    filepath = session.get('filepath')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'has_match': False})

    try:
        data_info = session.get('data_info', {})
        cols = [f.get('name', '') for f in data_info.get('fields', [])]
        match = find_match(
            filename=os.path.basename(filepath),
            file_size=os.path.getsize(filepath),
            original_columns=cols,
        )
        if match:
            return jsonify({
                'has_match': True,
                'filename': match.get('filename', ''),
                'updated_at': match.get('updated_at', ''),
            })
    except Exception as e:
        app.logger.error(f"[HISTORY] 检查映射历史失败: {e}")

    return jsonify({'has_match': False})


@app.route('/api/mapping-history/apply', methods=['POST'])
def apply_mapping_history():
    """API: 应用历史映射到当前会话"""
    filepath = session.get('filepath')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': '无上传文件'})

    try:
        data_info = session.get('data_info', {})
        cols = [f.get('name', '') for f in data_info.get('fields', [])]
        match = find_match(
            filename=os.path.basename(filepath),
            file_size=os.path.getsize(filepath),
            original_columns=cols,
        )
        if not match:
            return jsonify({'success': False, 'error': '未找到匹配的历史映射'})

        # 将历史映射写入 session
        session['field_mapping'] = match.get('field_mapping', {})
        session['mapping_prefilled'] = True

        balance_mapping = match.get('balance_field_mapping', {})
        if balance_mapping:
            session['balance_field_mapping'] = balance_mapping

        app.logger.info(f"[HISTORY] 已应用历史映射: {match.get('filename')}")
        return jsonify({'success': True})

    except Exception as e:
        app.logger.error(f"[HISTORY] 应用映射历史失败: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api-config', methods=['GET'])
def api_config_page():
    """API配置页面已废弃 — 直接跳转到字段映射"""
    if 'data_info' not in session:
        return redirect(url_for('upload_page'))
    return redirect(url_for('field_mapper_page'))

@app.route('/query', methods=['GET'])
def query_page():
    """查询页面"""
    if 'data_info' not in session:
        return redirect(url_for('upload_page'))

    # 传递数据信息到模板
    data_info = session.get('data_info', {})
    # Dify 复核模型始终可用（硬编码配置）
    return render_template('query.html', data_info=data_info, has_review_config=True)

# 注释掉results页面，因为查询页面已经包含完整的结果展示功能
# @app.route('/results', methods=['GET'])
# def results_page():
#     """结果页面"""
#     if 'data_info' not in session:
#         return redirect(url_for('upload_page'))
#     return render_template('results.html')

# ── 财务报表清洗路由 ──

@app.route('/report-cleaner', methods=['GET'])
def report_cleaner_page():
    """财务报表清洗页面 - 已迁移至完整性测试页面"""
    return redirect(url_for('integrity_test_page'))


@app.route('/api/report-upload', methods=['POST'])
def api_report_upload():
    """上传报表文件并预览"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未选择文件'})

    file = request.files['file']
    if not file.filename:
        return jsonify({'success': False, 'error': '文件名为空'})

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return jsonify({'success': False, 'error': '仅支持 .xlsx 和 .xls 格式'})

    try:
        from modules.report_cleaner import ReportCleaner
        fname = f"report_{int(time.time())}_{secure_filename(file.filename)}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], fname)
        file.save(filepath)

        cleaner = ReportCleaner()
        sheets = cleaner.load_file(filepath)

        # 存 session
        session['report_filepath'] = filepath
        session.modified = True

        return jsonify({
            'success': True,
            'filepath': filepath,
            'filename': file.filename,
            'sheets': sheets,
        })
    except Exception as e:
        app.logger.error(f"[REPORT-UPLOAD] {e}")
        return jsonify({'success': False, 'error': f'文件读取失败: {e}'})


@app.route('/api/report-detect', methods=['POST'])
def api_report_detect():
    """AI 检测报表结构"""
    data = request.get_json()
    sheet_name = data.get('sheet_name', '')
    filepath = session.get('report_filepath')

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': '请先上传文件'})

    try:
        dify = _get_dify_client()
        from modules.report_cleaner import ReportCleaner
        cleaner = ReportCleaner()
        cleaner.load_file(filepath)

        meta = cleaner.ai_detect(sheet_name, dify)
        return jsonify(meta)
    except Exception as e:
        app.logger.error(f"[REPORT-DETECT] {e}")
        return jsonify({'success': False, 'error': f'AI 检测失败: {e}'})


@app.route('/api/report-extract', methods=['POST'])
def api_report_extract():
    """规则提取清洗数据"""
    data = request.get_json()
    sheet_name = data.get('sheet_name', '')
    detection_meta = data.get('detection_meta', {})
    filepath = session.get('report_filepath')

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': '请先上传文件'})

    try:
        from modules.report_cleaner import ReportCleaner
        cleaner = ReportCleaner()
        cleaner.load_file(filepath)
        result = cleaner.extract_by_meta(sheet_name, detection_meta)

        if result.get('success'):
            session['report_clean_data'] = result
            session.modified = True

        return jsonify(result)
    except Exception as e:
        app.logger.error(f"[REPORT-EXTRACT] {e}")
        return jsonify({'success': False, 'error': f'提取失败: {e}'})


@app.route('/api/report-export', methods=['POST'])
def api_report_export():
    """导出清洗后数据为 Excel"""
    data = request.get_json()
    sheet_name = data.get('sheet_name', '')
    detection_meta = data.get('detection_meta', {})
    filepath = session.get('report_filepath')

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': '请先上传文件'})

    try:
        from modules.report_cleaner import ReportCleaner
        cleaner = ReportCleaner()
        cleaner.load_file(filepath)
        export_data = cleaner.export_to_excel(sheet_name, detection_meta)

        return send_file(
            export_data,
            as_attachment=True,
            download_name=f'报表清洗_{os.path.basename(filepath)}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        app.logger.error(f"[REPORT-EXPORT] {e}")
        return jsonify({'success': False, 'error': f'导出失败: {e}'})

@app.route('/api/report-reconciliation', methods=['POST'])
def api_report_reconciliation():
    """科目余额表 ↔ 报表期末余额核对"""
    data = request.get_json()
    sheet_name = data.get('sheet_name', '')
    detection_meta = data.get('detection_meta', {})
    filepath = session.get('report_filepath')

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': '请先上传文件'})

    # 构建 Dify 客户端（给 AI 兜底用，可选）
    dify_client = None
    try:
        dify_client = _get_dify_client()
    except Exception:
        pass  # 没有 Dify 配置时不阻断，只跳过 AI 兜底

    try:
        from modules.report_reconciliation import ReconciliationEngine
        from modules.report_cleaner import ReportCleaner

        # 提取/合并清洗后的报表数据（支持多报表：前端直接传 report_data_list）
        report_data_list = data.get('report_data_list')
        if report_data_list and len(report_data_list) > 0:
            # 前端已提取好，合并所有报表数据
            combined_rows = []
            for rd in report_data_list:
                if rd.get('data'):
                    combined_rows.extend(rd['data'])
            report_data = {
                'success': True,
                'data': combined_rows,
                'columns': report_data_list[0].get('columns', []),
                'report_type': 'combined',
                'row_count': len(combined_rows),
            }
        else:
            # 旧逻辑：从文件提取
            cleaner = ReportCleaner()
            cleaner.load_file(filepath)
            report_data = cleaner.extract_by_meta(sheet_name, detection_meta)
            if not report_data.get("success"):
                return jsonify(report_data)

        # 科目余额表字段定义
        balance_info = session.get('balance_data_info', {})
        balance_fields = balance_info.get('mapped_fields', balance_info.get('fields', []))

        # DuckDB 连接 — 优先用完整性测试快照，其次末级科目版，兜底原始数据
        engine = get_duckdb_engine()
        if engine.table_exists('balance_integrity'):
            balance_table = 'balance_integrity'
        elif engine.table_exists('balance_data'):
            balance_table = 'balance_data'
        else:
            return jsonify({'success': False, 'error': '未找到科目余额表数据，请先导入科目余额表'})

        cursor = engine._conn

        # 调试：检查数据
        debug_info = {}
        try:
            sample = cursor.execute(
                f'SELECT * FROM "{balance_table}" LIMIT 3'
            ).fetchall()
            cols = [d[0] for d in cursor.description]
            debug_info['table'] = balance_table
            debug_info['columns'] = cols
            debug_info['sample'] = [dict(zip(cols, r)) for r in sample]
        except Exception as e:
            debug_info['error'] = str(e)

        reconciler = ReconciliationEngine(
            db_cursor=cursor,
            balance_fields=balance_fields,
            balance_table=balance_table,
        )

        # 判断是初始映射还是刷新核对
        mappings = data.get('mappings')
        if mappings:
            # 用户修改了映射 → 重新核对
            result = reconciler.reconcile_with_mappings(mappings, report_data)
        else:
            # 初始调用 → 返回映射数据
            result = reconciler.get_balance_mappings(
                report_data,
                dify_client=dify_client,
            )
        # 把引擎实际使用的字段名也加到调试信息中
        if isinstance(result, dict) and '_fields' in result:
            debug_info['engine_fields'] = result.pop('_fields')
        result['_debug'] = debug_info
        return jsonify(result)

    except Exception as e:
        app.logger.error(f"[REPORT-RECONCILE] {e}")
        return jsonify({'success': False, 'error': f'核对失败: {e}'})

@app.route('/api/report-reconciliation/export', methods=['POST'])
def api_report_reconciliation_export():
    """导出核对结果到 Excel"""
    data = request.get_json()
    mappings = data.get('mappings', [])
    sheet_name = data.get('sheet_name', '')
    detection_meta = data.get('detection_meta', {})
    filepath = session.get('report_filepath')

    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': '请先上传文件'})

    try:
        from modules.report_reconciliation import ReconciliationEngine
        from modules.report_cleaner import ReportCleaner
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # 提取/合并报表数据
        report_data_list = data.get('report_data_list')
        if report_data_list and len(report_data_list) > 0:
            combined_rows = []
            for rd in report_data_list:
                if rd.get('data'):
                    combined_rows.extend(rd['data'])
            report_data = {
                'success': True,
                'data': combined_rows,
                'columns': report_data_list[0].get('columns', []),
                'report_type': 'combined',
                'row_count': len(combined_rows),
            }
        else:
            cleaner = ReportCleaner()
            cleaner.load_file(filepath)
            report_data = cleaner.extract_by_meta(sheet_name, detection_meta)
            if not report_data.get("success"):
                return jsonify(report_data)

        # 执行核对
        engine = get_duckdb_engine()
        balance_info = session.get('balance_data_info', {})
        balance_fields = balance_info.get('mapped_fields', balance_info.get('fields', []))
        balance_table = 'balance_integrity' if engine.table_exists('balance_integrity') else 'balance_data'

        reconciler = ReconciliationEngine(
            db_cursor=engine._conn,
            balance_fields=balance_fields,
            balance_table=balance_table,
        )
        result = reconciler.reconcile_with_mappings(mappings, report_data)
        if not result.get("success"):
            return jsonify(result)

        comparison = result.get("comparison", [])
        stats = result.get("stats", {})

        # 生成 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "核对结果"

        hfont = Font(bold=True, size=11, color="FFFFFF")
        hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        h_align = Alignment(horizontal="center", vertical="center")
        bborder = Border(bottom=Side(style='thin', color='d9d9d9'))

        # 标题行
        ws.cell(row=1, column=1, value="科目余额表 ↔ 报表核对结果").font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        ws.cell(row=2, column=1, value=f"报表项目: {stats.get('total_items', 0)} | 已一致: {stats.get('matched', 0)} | 有差异: {stats.get('difference', 0)} | 匹配率: {stats.get('match_rate', 0)}%").font = Font(size=10, color='666666')
        ws.merge_cells('A2:E2')

        headers = ["报表项目", "报表金额", "余额表汇总", "差异", "匹配科目"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=ci, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = h_align

        diff_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
        row_idx = 5
        for item in comparison:
            is_diff = abs(item.get("diff", 0)) > 0.01
            accounts = item.get("matched_accounts", [])
            detail = "; ".join(f"{a.get('name','')}" for a in accounts[:10]) if accounts else ""
            if item["match_type"] == "report_only":
                detail = "仅报表有"
            elif item["match_type"] == "unmatched":
                detail = "未匹配科目"

            vals = [
                item.get("report_item", ""),
                item.get("report_amount", 0),
                item.get("balance_amount", 0),
                round(item.get("diff", 0), 2),
                detail,
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=ci, value=v)
                cell.border = bborder
                if is_diff:
                    cell.fill = diff_fill
                if ci in (2, 3, 4) and isinstance(v, (int, float)):
                    cell.number_format = '#,##0.00'
            row_idx += 1

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 40

        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(
            buf,
            as_attachment=True,
            download_name="核对结果.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        app.logger.error(f"[RECONCILE-EXPORT] {e}")
        return jsonify({'success': False, 'error': f'导出失败: {e}'})

@app.route('/api/report-reconciliation/ai-analyze', methods=['POST'])
def api_report_reconciliation_ai_analyze():
    """AI 差异分析：分析核对结果的差异模式并给出建议"""
    data = request.get_json()
    comparison = data.get('comparison', [])
    mappings = data.get('mappings', [])

    # 构造差异数据
    diff_lines = []
    for c in comparison:
        diff = abs(c.get("diff", 0))
        if diff > 0.01:
            accounts = c.get("matched_accounts", [])
            acct_details = "; ".join(f"{a.get('code','')} {a.get('name','')}" for a in accounts[:5])
            diff_lines.append(
                f"- {c.get('report_item','')}: "
                f"报表={c.get('report_amount',0):.2f}, "
                f"余额表={c.get('balance_amount',0):.2f}, "
                f"差异={c.get('diff',0):.2f}"
                f"{'  匹配科目: ' + acct_details if acct_details else ''}"
            )

    # 科目映射样例
    sample_mappings = mappings[:30]
    map_lines = []
    acct_groups = {}
    for m in sample_mappings:
        ri = m.get('report_item', '') or '(未映射)'
        if ri not in acct_groups:
            acct_groups[ri] = []
        acct_groups[ri].append(f"{m.get('account_code','')} {m.get('account_name','')}")

    for ri, accts in sorted(acct_groups.items()):
        map_lines.append(f"  {ri}: {', '.join(accts[:8])}")

    prompt = f"""你是一个审计数据核对专家。请分析以下科目余额表与财务报表的核对差异，找出可能的映射错误并提出建议。

## 当前科目→报表映射关系（部分）
{chr(10).join(map_lines)}

## 存在差异的项目
{chr(10).join(diff_lines) if diff_lines else '(无差异)'}

## 要求
请分析差异中是否存在以下模式，针对每种发现给出具体建议：

1. **映射调换**: 两个项目的差异金额接近，可能是科目映射反了（如 A 科目应映射到 B 项目，B 科目应映射到 A 项目）
2. **归属错误**: 某个科目的余额可能被归到了错误的报表项目下
3. **遗漏科目**: 报表中有金额但余额表没有对应科目的项目
4. **其他异常**: 你注意到的其他差异特征

请用中文回答，每条建议需指明可能的科目和报表项目名称。
"""

    try:
        # 通过 Dify 代理调用 AI
        dify = _get_dify_client()
        analysis = dify.chat(
            "你是一个经验丰富的审计数据核对专家。",
            prompt,
            timeout=30,
        )
        return jsonify({'success': True, 'analysis': analysis})

    except Exception as e:
        app.logger.error(f"[RECONCILE-AI] 分析失败: {e}")
        return jsonify({'success': False, 'error': f'分析失败: {e}'})


@app.route('/api/upload/preview', methods=['POST'])
def upload_file_preview():
    """API: 上传文件并预览结构（检测 sheet 和原始行），不进行数据解析"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '没有选择文件'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '没有选择文件'})

    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': '不支持的文件类型'})

    # 清理旧的临时文件
    old_pending = session.pop('pending_filepath', None)
    if old_pending and os.path.exists(old_pending):
        try:
            os.remove(old_pending)
        except Exception:
            pass

    # 保存文件到临时路径
    temp_filename = f"temp_{int(time.time())}_{file.filename}"
    temp_path = os.path.join(app.config['UPLOAD_FOLDER'], temp_filename)
    file.save(temp_path)

    ext = os.path.splitext(file.filename)[1].lower()
    is_excel = ext == '.xlsx'

    try:
        sheets = {}
        sheet_names = []

        if is_excel:
            # 检测 sheet 名称
            sheet_names = DataProcessor.get_xlsx_sheet_names(temp_path)
            # 获取原始预览
            raw = DataProcessor.preview_raw(temp_path, nrows=15)
            for name in sheet_names:
                info = raw.get(name, {})
                sheets[name] = {
                    'total_rows': info.get('total_rows', 0),
                    'total_cols': info.get('total_cols', 0),
                    'preview_rows': info.get('rows', []),
                }
        else:
            # CSV 当作单 sheet
            raw = DataProcessor.preview_raw(temp_path, nrows=15)
            csv_info = raw.get('__csv__', {})
            sheet_names = ['__csv__']
            sheets['__csv__'] = {
                'total_rows': csv_info.get('total_rows', 0),
                'total_cols': csv_info.get('total_cols', 0),
                'preview_rows': csv_info.get('rows', []),
            }

        # 保存文件路径到 session 供后续解析使用
        session['pending_filepath'] = temp_path
        session['pending_filename'] = file.filename

        return jsonify({
            'success': True,
            'filename': file.filename,
            'file_type': 'excel' if is_excel else 'csv',
            'sheet_names': sheet_names,
            'sheets': sheets,
            'is_multi_sheet': len(sheet_names) > 1,
        })

    except Exception as e:
        import traceback
        app.logger.error(f"[UPLOAD PREVIEW ERROR] {str(e)}\n{traceback.format_exc()}")
        # 清理临时文件
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass
        return jsonify({'success': False, 'error': f'文件预览失败: {str(e)}'})


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """API: 处理文件上传（支持 sheet 选择和 header row 指定）"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '没有选择文件'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '没有选择文件'})

    # 检查文件扩展名
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': '不支持的文件类型'})

    # 清理旧的 DuckDB 数据（上传新文件前）
    cleanup_session_db()
    session.pop('data_info', None)
    session.pop('filepath', None)
    session.pop('field_mapping', None)
    session.pop('balance_field_mapping', None)

    # 保存文件
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filepath)

    # 解析可选参数：sheet_name 和 header_row
    sheet_name = request.form.get('sheet_name') or None
    header_row_raw = request.form.get('header_row')
    header_row = int(header_row_raw) - 1 if header_row_raw and header_row_raw.isdigit() else None
    # header_row 在前端从 1 开始计数，转为 0-indexed 传给 pandas

    app.logger.info(f"[UPLOAD] sheet_name={sheet_name}, header_row={header_row}(前端输入={header_row_raw})")

    # 调用DataProcessor处理文件
    try:
        processor = DataProcessor(filepath)
        data_info = processor.process(
            sheet_name=sheet_name if sheet_name != '__csv__' else None,
            header_row=header_row,
        )
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"文件处理失败详情: {error_details}")
        # 清理上传失败的文件
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
                app.logger.info(f"[CLEANUP] 已删除上传失败的文件: {filepath}")
            except Exception as cleanup_err:
                app.logger.error(f"[CLEANUP] 文件删除失败: {cleanup_err}")
        return jsonify({'success': False, 'error': f'文件处理失败: {str(e)}'})

    # 存储到会话
    session['filepath'] = filepath
    session['data_info'] = data_info
    session['upload_options'] = {'sheet_name': sheet_name, 'header_row': header_row}
    session['duckdb_imported'] = False

    return jsonify(data_info)


@app.route('/api/review-code', methods=['POST'])
def review_code():
    """API: 复核 AI 生成的 SQL 代码 — 由第二 AI 模型（或同级模型）从语法、安全、意图、性能四个维度审查"""
    data = request.json
    code = data.get('code', '')
    query = data.get('query', '')

    if not code:
        return jsonify({'success': False, 'error': '代码不能为空'})

    try:
        # 获取字段信息用于上下文
        data_info = session.get('data_info', {})
        fields = data_info.get('mapped_fields', data_info.get('fields', []))
        fields_desc = "\n".join([f"- {f['name']} ({f['type']})" for f in fields[:20]])

        prompt = f"""你是一名资深的财务数据 SQL 审查专家。请从以下四个维度审查 AI 生成的 DuckDB SQL 代码。

## 用户查询
{query}

## 生成的 SQL 代码
```sql
{code}
```

## 数据表字段
{fields_desc or '（无详细字段信息）'}

## 审查维度
1. **语法检查** — 是否符合 DuckDB SQL 语法？
2. **安全审查** — 是否仅包含 SELECT 只读操作？有无危险语句？
3. **意图匹配** — SQL 逻辑是否准确反映了用户的查询需求？
4. **性能优化** — 是否有明显的性能问题？能否优化？

## 输出格式
请输出以下 JSON 格式，注意必须是合法的 JSON，不要包含 Markdown 代码块包裹：
{{
  "passed": true/false,
  "summary": "一句话总结审查结论",
  "aspects": [
    {{"name": "语法检查", "passed": true/false, "reason": "详细说明"}},
    {{"name": "安全审查", "passed": true/false, "reason": "详细说明"}},
    {{"name": "意图匹配", "passed": true/false, "reason": "详细说明"}},
    {{"name": "性能优化", "passed": true/false, "reason": "详细说明"}}
  ]
}}"""

        # 通过复核 Dify Workflow 调用 AI
        dify = _get_review_dify_client()
        content = dify.chat(
            "你是一名资深的 SQL 审查专家。只输出 JSON，不要加 Markdown 代码块包裹。",
            prompt,
            timeout=60,
        )

        # 清理 Markdown 包裹
        content = content.strip()
        if content.startswith('```'):
            content = re.sub(r'^```(?:json)?\s*', '', content)
            content = re.sub(r'\s*```$', '', content)
            content = content.strip()

        # 提取 JSON 对象
        brace_start = content.find('{')
        brace_end = content.rfind('}')
        if brace_start >= 0 and brace_end >= 0:
            content = content[brace_start:brace_end + 1]

        review_result = json_module.loads(content)

        return jsonify({
            'success': True,
            'review': review_result,
            'model_used': 'Dify (Qwen3-30B-A3B)',
        })

    except Exception as e:
        import traceback
        app.logger.error(f"[REVIEW-CODE] {str(e)}\n{traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': f'代码复核失败: {str(e)}',
        })


@app.route('/api/generate-code', methods=['POST'])
def generate_code():
    """API: 生成SQL查询"""
    data = request.json
    query = data.get('query')

    if not query:
        return jsonify({'success': False, 'error': '查询语句不能为空'})

    data_info = session.get('data_info')
    if not data_info:
        return jsonify({'success': False, 'error': '请先上传数据文件'})

    try:
        # 使用 Dify 代理
        dify = _get_dify_client()
        generator = AICodeGenerator(dify_client=dify)
        fields = data_info.get('mapped_fields', data_info.get('fields', []))
        preview = data_info.get('mapped_preview', data_info.get('preview', []))

        # 如果已导入 DuckDB，使用数据库 schema 信息
        if session.get('duckdb_imported'):
            engine = get_duckdb_engine()
            fields = engine.get_schema('data')
            app.logger.info(f"[GENERATE_CODE] Using DuckDB schema: {fields}")

        app.logger.info(f"[GENERATE_CODE] data_info keys: {list(data_info.keys())}")
        sql = generator.generate(query, fields, preview)
        log_generate(
            session_id=session.get('session_id', ''),
            user_query=query,
            generated_sql=sql,
            success=True,
        )
        # 保存查询历史到 session
        history = session.get('query_history', [])
        history.append({
            'query': query,
            'sql': sql,
            'timestamp': time.strftime('%H:%M:%S'),
            'success': True,
        })
        session['query_history'] = history[-50:]  # 最多保留 50 条
        # 保存当前查询信息供导出使用
        session['last_query_info'] = {
            'query_text': query,
            'sql_code': sql,
        }
        # 自动生成代码解释供导出使用
        try:
            session['last_explanation'] = generator.explain_code(sql)
        except Exception:
            session['last_explanation'] = ''
        return jsonify({'success': True, 'code': sql})
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        app.logger.error(f"[GENERATE_CODE ERROR] {str(e)}\n{tb}")
        log_generate(
            session_id=session.get('session_id', ''),
            user_query=query,
            generated_sql='',
            success=False,
            error=str(e),
        )
        history = session.get('query_history', [])
        history.append({
            'query': query,
            'sql': '',
            'timestamp': time.strftime('%H:%M:%S'),
            'success': False,
        })
        session['query_history'] = history[-50:]
        return jsonify({'success': False, 'error': f'SQL生成失败: {str(e)}'})


@app.route('/api/query-history', methods=['GET'])
def query_history():
    """API: 返回当前会话的查询历史"""
    history = session.get('query_history', [])
    return jsonify({'success': True, 'history': history})

@app.route('/api/saved-queries', methods=['GET', 'POST'])
def saved_queries():
    """API: 收藏查询的列表/保存"""
    if request.method == 'GET':
        saved = session.get('saved_queries', [])
        summary = []
        for q in saved:
            summary.append({
                'id': q['id'],
                'name': q['name'],
                'query': q['query'],
                'timestamp': q['timestamp'],
                'row_count': q.get('row_count'),
            })
        return jsonify({'success': True, 'queries': summary})

    # POST - 保存
    data = request.json
    name = data.get('name', '未命名查询')
    query = data.get('query', '')
    sql = data.get('sql', '')
    result_data = data.get('result_data')

    if not query:
        return jsonify({'success': False, 'error': '查询语句不能为空'})

    import uuid
    qid = str(uuid.uuid4())[:8]

    row_count = None
    if result_data:
        rd = result_data.get('result', result_data)
        if rd.get('type') == 'dataframe':
            row_count = len(rd.get('data', []))

    saved = session.get('saved_queries', [])

    result_file = None
    if result_data:
        try:
            result_filename = f"saved_{qid}_{int(time.time())}.json"
            result_filepath = os.path.join(app.config['UPLOAD_FOLDER'], result_filename)
            with open(result_filepath, 'w', encoding='utf-8') as f:
                json_module.dump(_json_safe(result_data), f, ensure_ascii=False)
            result_file = result_filename
        except Exception as e:
            app.logger.warning(f"[SAVED_QUERY] 保存结果文件失败: {e}")

    entry = {
        'id': qid,
        'name': name,
        'query': query,
        'sql': sql,
        'timestamp': time.strftime('%Y-%m-%d %H:%M'),
        'row_count': row_count,
        'result_file': result_file,
    }
    saved.append(entry)
    session['saved_queries'] = saved
    return jsonify({'success': True, 'id': qid})


@app.route('/api/saved-queries/<qid>', methods=['DELETE'])
def delete_saved_query(qid):
    """API: 删除收藏查询"""
    saved = session.get('saved_queries', [])
    for i, q in enumerate(saved):
        if q['id'] == qid:
            if q.get('result_file'):
                fpath = os.path.join(app.config['UPLOAD_FOLDER'], q['result_file'])
                if os.path.exists(fpath):
                    try:
                        os.remove(fpath)
                    except Exception:
                        pass
            saved.pop(i)
            session['saved_queries'] = saved
            return jsonify({'success': True})
    return jsonify({'success': False, 'error': '未找到该收藏查询'})


@app.route('/api/export-saved-queries', methods=['POST'])
def export_saved_queries():
    """API: 批量导出收藏查询为多 Sheet Excel"""
    data = request.json
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'error': '请选择要导出的查询'})

    saved = session.get('saved_queries', [])
    selected = [q for q in saved if q['id'] in ids]
    if not selected:
        return jsonify({'success': False, 'error': '未找到选中的查询'})

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        import io

        filename = '收藏查询汇总.xlsx'
        wb = openpyxl.Workbook()
        # remove default sheet — we create per-query sheets below
        wb.remove(wb.active)
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        for idx, sq in enumerate(selected):
            sheet_name = sq.get('name', f'查询{idx+1}')[:31]
            ws = wb.create_sheet(sheet_name)

            result_data = None
            if sq.get('result_file'):
                rfpath = os.path.join(app.config['UPLOAD_FOLDER'], sq['result_file'])
                if os.path.exists(rfpath):
                    try:
                        with open(rfpath, 'r', encoding='utf-8') as f:
                            result_data = json_module.load(f)
                    except Exception:
                        pass

            if not result_data:
                ws.cell(row=1, column=1, value='项目').font = header_font
                ws.cell(row=1, column=1).fill = header_fill
                ws.cell(row=1, column=2, value='内容')
                ws.cell(row=1, column=2).font = header_font
                ws.cell(row=1, column=2).fill = header_fill
                ws.cell(row=2, column=1, value='查询名称')
                ws.cell(row=2, column=2, value=sq.get('name', ''))
                ws.cell(row=3, column=1, value='自然语言查询')
                ws.cell(row=3, column=2, value=sq.get('query', ''))
                ws.cell(row=4, column=1, value='SQL代码')
                ws.cell(row=4, column=2, value=sq.get('sql', ''))
                ws.cell(row=5, column=1, value='备注')
                ws.cell(row=5, column=2, value='结果数据不可用')
                ws.column_dimensions['A'].width = 16
                ws.column_dimensions['B'].width = 80
                continue

            rd = result_data.get('result', result_data)
            columns = rd.get('columns', [])
            rows = rd.get('data', [])

            # Info rows
            info_data = [
                ('查询名称', sq.get('name', '')),
                ('自然语言查询', sq.get('query', '')),
                ('SQL代码', sq.get('sql', '')),
                ('记录数', len(rows)),
            ]
            ws.cell(row=1, column=1, value='信息').font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=2, value='内容').font = header_font
            ws.cell(row=1, column=2).fill = header_fill
            for ri, (k, v) in enumerate(info_data, 2):
                ws.cell(row=ri, column=1, value=k)
                ws.cell(row=ri, column=2, value=v)
            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['B'].width = 80

            # Data rows starting from row 7
            if rows and columns:
                for ci, col_name in enumerate(columns, 1):
                    cell = ws.cell(row=7, column=ci, value=col_name)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
                for ri, row in enumerate(rows, 8):
                    for ci, col_name in enumerate(columns, 1):
                        val = row.get(col_name, '')
                        ws.cell(row=ri, column=ci, value=val if val is not None else '')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        app.logger.error(f"[EXPORT SAVED] 批量导出失败: {e}")
        return jsonify({'success': False, 'error': f'导出失败: {str(e)}'})





@app.route('/api/optimize-query', methods=['POST'])
def optimize_query():
    """API: 优化自然语言查询（本地词典 + AI 同义词扩展）"""
    data = request.json
    query = data.get('query')

    if not query:
        return jsonify({'success': False, 'error': '查询语句不能为空'})

    # 1. 本地词典初步扩展
    from modules.synonym_dict import expand_keywords, find_keywords
    local_expanded = expand_keywords(query)
    expanded_terms = [m['standard'] for m in find_keywords(query)]

    # 2. AI 二次扩展（通过 Dify 代理）
    try:
        dify = _get_dify_client()
        generator = AICodeGenerator(dify_client=dify)
        ai_optimized = generator.optimize_query(query, local_expanded=local_expanded)
        return jsonify({
            'success': True,
            'original_query': query,
            'local_expanded': local_expanded,
            'optimized_query': ai_optimized,
            'expanded_terms': expanded_terms,
            'source': 'ai'
        })
    except Exception as e:
        # AI 失败时降级到本地扩展
        return jsonify({
            'success': True,
            'original_query': query,
            'local_expanded': local_expanded,
            'optimized_query': local_expanded,
            'expanded_terms': expanded_terms,
            'source': 'local',
            'warning': f'AI 优化失败，已使用本地词典扩展: {str(e)}'
        })


@app.route('/api/explain-code', methods=['POST'])
def explain_code():
    """API: 解释代码功能"""
    data = request.json
    code = data.get('code')

    if not code:
        return jsonify({'success': False, 'error': '代码不能为空'})

    try:
        dify = _get_dify_client()
        generator = AICodeGenerator(dify_client=dify)
        explanation = generator.explain_code(code)
        session['last_explanation'] = explanation
        return jsonify({'success': True, 'explanation': explanation})
    except Exception as e:
        session['last_explanation'] = ''
        return jsonify({'success': False, 'error': f'代码解释失败: {str(e)}'})


@app.route('/api/preset-rules', methods=['GET'])
def preset_rules_list():
    """API: 返回所有预设筛选规则（扁平列表）"""
    rules = get_rules()
    return jsonify({'success': True, 'rules': rules})


@app.route('/api/preset-rules/packs', methods=['GET'])
def preset_rules_packs():
    """API: 返回规则包分组结构（含内置包 + 自定义规则）"""
    packs = get_packs()
    custom_rules = get_custom_rules()
    return jsonify({
        'success': True,
        'packs': packs,
        'custom_rules': custom_rules,
    })


@app.route('/api/preset-rules', methods=['POST'])
def preset_rules_save():
    """API: 保存自定义规则（新建或更新）"""
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': '请求数据不能为空'})
    if not data.get('name') or not data.get('sql_template'):
        return jsonify({'success': False, 'error': '规则名称和 SQL 模板不能为空'})
    rule = save_rule(data)
    return jsonify({'success': True, 'rule': rule})


@app.route('/api/preset-rules/<rule_id>', methods=['DELETE'])
def preset_rules_delete(rule_id):
    """API: 删除自定义规则"""
    ok = delete_rule(rule_id)
    if ok:
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': '仅允许删除自定义规则，或规则不存在'})


@app.route('/api/preset-rules/apply', methods=['POST'])
def preset_rules_apply():
    """API: 应用预设规则，返回生成的 SQL"""
    data = request.json
    rule_id = data.get('rule_id')
    params = data.get('params', {})
    if not rule_id:
        return jsonify({'success': False, 'error': 'rule_id 不能为空'})
    rule = get_rule_by_id(rule_id)
    if not rule:
        return jsonify({'success': False, 'error': f'未找到规则: {rule_id}'})
    sql = apply_rule(rule_id, params)
    if sql is None:
        return jsonify({'success': False, 'error': '规则应用失败，请检查参数'})
    return jsonify({
        'success': True,
        'sql': sql,
        'rule': {'id': rule['id'], 'name': rule['name'], 'category': rule['category']}
    })


@app.route('/api/sampling/methods', methods=['GET'])
def sampling_methods():
    """API: 返回所有取样方法"""
    return jsonify({'success': True, 'methods': get_methods()})


@app.route('/api/sampling/execute', methods=['POST'])
def sampling_execute():
    """API: 执行取样，返回结果"""
    data = request.json
    method_id = data.get('method_id')
    params = data.get('params', {})

    if not method_id:
        return jsonify({'success': False, 'error': 'method_id 不能为空'})

    methods = get_methods()
    method = next((m for m in methods if m['id'] == method_id), None)
    if not method:
        return jsonify({'success': False, 'error': f'未找到取样方法: {method_id}'})

    sql = generate_sql(method_id, params)
    if sql is None:
        return jsonify({'success': False, 'error': '无法生成取样 SQL'})

    engine = get_duckdb_engine()
    if not engine.table_exists('data'):
        return jsonify({'success': False, 'error': '数据尚未导入，请先完成字段映射'})

    result = engine.execute(sql)
    if result.get('success'):
        return jsonify({
            'success': True,
            'method': method['name'],
            'result': result.get('result', {}),
            'sql': sql.strip()
        })
    else:
        return jsonify({'success': False, 'error': f'取样执行失败: {result.get("error")}'})


@app.route('/api/execute', methods=['POST'])
def execute_code():
    """API: 执行 SQL 查询"""
    data = request.json
    sql = data.get('code')

    if not sql:
        return jsonify({'success': False, 'error': 'SQL不能为空'})

    engine = get_duckdb_engine()

    # 优先通过 session 标记快速判断，标记丢失时检查实际的数据库状态
    if not session.get('duckdb_imported'):
        if not engine.table_exists('data'):
            # 表不存在，尝试重新导入
            filepath = session.get('filepath')
            field_mapping = session.get('field_mapping')
            upload_opts = session.get('upload_options', {})
            sheet_name = upload_opts.get('sheet_name')
            header_row = upload_opts.get('header_row')
            constant_cols = session.get('manual_fills') or None
            if filepath and field_mapping:
                try:
                    ext = os.path.splitext(filepath)[1].lower()
                    reverse_mapping = {v: k for k, v in field_mapping.items()}
                    if ext == '.csv':
                        engine.import_csv(filepath, 'data', reverse_mapping, header_row=header_row,
                                          constant_columns=constant_cols)
                    elif ext == '.xlsx':
                        engine.import_xlsx(filepath, 'data', reverse_mapping,
                                           sheet_name=sheet_name, header_row=header_row,
                                           constant_columns=constant_cols)
                    else:
                        return jsonify({'success': False, 'error': '不支持的文件格式，请重新上传'})
                    session['duckdb_imported'] = True
                    app.logger.info(f"[DUCKDB] 执行查询前自动重导入成功: {filepath}")
                except Exception as e:
                    app.logger.error(f"[DUCKDB] 执行查询前重导入失败: {e}")
                    return jsonify({'success': False, 'error': f'数据尚未导入DuckDB，请先配置字段映射'})
            else:
                return jsonify({'success': False, 'error': '数据尚未导入DuckDB，请先配置字段映射'})
        else:
            # 表存在但标记丢失（如调试重启），修复标记
            session['duckdb_imported'] = True

    try:
        result = engine.execute(sql)

        # 审计日志
        row_count = None
        if result.get('success') and result.get('result', {}).get('type') == 'dataframe':
            row_count = len(result['result'].get('data', []))
        log_execute(
            session_id=session.get('session_id', ''),
            sql=sql,
            success=result.get('success', False),
            row_count=row_count,
            execution_time=result.get('execution_time'),
            error=None if result.get('success') else result.get('error'),
        )

        session['last_execution_result'] = {
            'success': result.get('success', False),
            'simplified_result': result.get('result', {})
        }
        # 保存实际执行的 SQL（可能被用户编辑过）
        info = session.get('last_query_info', {'query_text': '', 'sql_code': ''})
        info['sql_code'] = sql
        session['last_query_info'] = info

        return jsonify(result)
    except Exception as e:
        import traceback
        app.logger.error(f"[DUCKDB EXECUTE ERROR] {str(e)}\n{traceback.format_exc()}")
        log_execute(
            session_id=session.get('session_id', ''),
            sql=sql,
            success=False,
            error=str(e),
        )
        return jsonify({'success': False, 'error': f'SQL执行失败: {str(e)}'})

@app.route('/api/export', methods=['POST'])
def export_data():
    """API: 导出数据"""
    data = request.json
    format_type = data.get('format', 'csv')

    # 获取上次执行结果
    execution_result = session.get('last_execution_result')

    if not execution_result or not execution_result.get('success'):
        return jsonify({'success': False, 'error': '没有可导出的执行结果，请先执行查询'})

    # 检查是否有存储在临时文件中的完整结果
    result_data = None
    simplified_result = execution_result.get('simplified_result')
    if simplified_result and simplified_result.get('result_file'):
        # 从临时文件加载完整结果
        result_filename = simplified_result['result_file']
        result_filepath = os.path.join(app.config['UPLOAD_FOLDER'], result_filename)
        if os.path.exists(result_filepath):
            try:
                import json as json_module
                with open(result_filepath, 'r', encoding='utf-8') as f:
                    result_data = json_module.load(f)
            except Exception as e:
                app.logger.error(f"加载结果文件失败: {e}")
                # 继续尝试使用简化结果
                pass

    # 如果没有临时文件，尝试从会话中的结果获取
    if result_data is None:
        result_data = execution_result.get('result')
    if result_data is None:
        result_data = execution_result.get('simplified_result')
    if not result_data:
        return jsonify({'success': False, 'error': '执行结果为空，无法导出'})

    try:
        # 生成文件名
        import time
        timestamp = int(time.time())
        filename = f'export_{timestamp}.{format_type}'
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        # ---- CSV 导出：优先用 DuckDB COPY（原生写磁盘，比 Python 快数量级） ----
        if format_type == 'csv':
            sql_code = session.get('last_query_info', {}).get('sql_code', '')
            engine = get_duckdb_engine()
            if sql_code and session.get('duckdb_imported') and engine.table_exists('data'):
                try:
                    # DuckDB COPY 直接写 CSV，绕过 Python 内存
                    engine._conn.execute(f"COPY ({sql_code}) TO '{filepath}' (HEADER, DELIMITER ',')")
                    return send_file(filepath, as_attachment=True, download_name=filename)
                except Exception as e:
                    app.logger.warning(f"[EXPORT] DuckDB COPY 失败，回退到 csv.writer: {e}")

            # 回退：从 session 结果写 CSV（流式写入，也够快）
            import csv
            result_type = result_data.get('type')
            if result_type != 'dataframe':
                return jsonify({'success': False, 'error': '不支持此类型导出为 CSV'})
            headers = result_data.get('columns', [])
            data_rows = result_data.get('data', [])
            if not data_rows:
                return jsonify({'success': False, 'error': '没有可导出的数据'})
            with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for row in data_rows:
                    writer.writerow([row.get(h, '') for h in headers])
            return send_file(filepath, as_attachment=True, download_name=filename)

        # ---- XLSX 导出 ----
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        # 归一化数据
        headers = []
        data_rows = []
        result_type = result_data.get('type')
        if result_type == 'dataframe':
            headers = result_data.get('columns', [])
            data_rows = result_data.get('data', [])
        elif result_type in ('list', 'tuple'):
            values = result_data.get('values', [])
            if values and isinstance(values[0], dict):
                headers = list(values[0].keys())
                data_rows = values
            else:
                headers = ['值']
                data_rows = [{'值': v} for v in values]
        elif result_type == 'dict':
            dd = result_data.get('data', {})
            if dd:
                headers = list(dd.keys())
                data_rows = [dd]
        elif result_type == 'scalar':
            data_rows = [{'结果': result_data.get('value')}]
            headers = ['结果']
        else:
            return jsonify({'success': False, 'error': f'不支持的结果类型: {result_type}'})
        if not data_rows:
            return jsonify({'success': False, 'error': '没有可导出的数据'})

        query_info = session.get('last_query_info', {})
        explanation = session.get('last_explanation', '')
        hfont = Font(bold=True, color='FFFFFF', size=11)
        hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        # 大数据量使用 write_only 模式（流式写入，不驻留内存）
        is_large = len(data_rows) > 10000
        if is_large:
            wb = openpyxl.Workbook(write_only=True)
            # Sheet 1: 查询信息（先用常规模式写好 info sheet）
            # write_only 下不能改 sheet 名/样式，直接用一个常规 workbook 写 info
            import io as _io
            info_wb = openpyxl.Workbook()
            info_ws = info_wb.active
            info_ws.title = '查询信息'
            for ci, val in enumerate(['项目', '内容'], 1):
                c = info_ws.cell(row=1, column=ci, value=val)
                c.font = hfont
                c.fill = hfill
            for ri, (k, v) in enumerate([('自然语言查询', query_info.get('query_text', '')),
                                          ('SQL代码', query_info.get('sql_code', '')),
                                          ('代码解释', explanation)], 2):
                info_ws.cell(row=ri, column=1, value=k)
                info_ws.cell(row=ri, column=2, value=v)
            info_ws.column_dimensions['A'].width = 14
            info_ws.column_dimensions['B'].width = 90
            info_ws.row_dimensions[2].height = 40
            info_ws.row_dimensions[3].height = 120
            info_ws.row_dimensions[4].height = 200
            # Sheet 2: 查询结果（write_only 流式写入）
            ws2 = wb.create_sheet(title='查询结果')
            ws2.append(headers)
            for row in data_rows:
                ws2.append([row.get(h, '') if row.get(h) is not None else '' for h in headers])
            # 合并两个 workbook
            wb.save(filepath)
            # 用 info workbook 覆盖写入同名 sheet（openpyxl 会替换已有 sheet）
            from openpyxl import load_workbook
            wb2 = load_workbook(filepath)
            # 删除 write_only 版自动生成的 info sheet（如果有）
            if '查询信息' in wb2.sheetnames:
                del wb2['查询信息']
            # 插入 info 作为第一个 sheet
            wb2._sheets.insert(0, info_ws._workspace if hasattr(info_ws, '_workspace') else None)
            # 更简单：直接复制 info_ws 内容
            wb2.create_sheet('查询信息', 0)
            ws_dest = wb2['查询信息']
            for row in info_ws.iter_rows(min_row=1, max_row=info_ws.max_row, max_col=info_ws.max_column, values_only=False):
                vals = [cell.value for cell in row]
                ws_dest.append(vals)
                # copy styles for header
                if row[0].row == 1:
                    for ci in range(1, len(vals)+1):
                        ws_dest.cell(row=1, column=ci).font = hfont
                        ws_dest.cell(row=1, column=ci).fill = hfill
            wb2.save(filepath)
        else:
            # 小数据量：常规模式，带样式
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = '查询信息'
            ws1.cell(row=1, column=1, value='项目').font = hfont
            ws1.cell(row=1, column=1).fill = hfill
            ws1.cell(row=1, column=2, value='内容').font = hfont
            ws1.cell(row=1, column=2).fill = hfill
            for ri, (k, v) in enumerate([('自然语言查询', query_info.get('query_text', '')),
                                          ('SQL代码', query_info.get('sql_code', '')),
                                          ('代码解释', explanation)], 2):
                ws1.cell(row=ri, column=1, value=k)
                ws1.cell(row=ri, column=2, value=v)
            ws1.column_dimensions['A'].width = 14
            ws1.column_dimensions['B'].width = 90
            ws1.row_dimensions[2].height = 40
            ws1.row_dimensions[3].height = 120
            ws1.row_dimensions[4].height = 200
            # 数据 sheet
            ws2 = wb.create_sheet('查询结果')
            for ci, h in enumerate(headers, 1):
                ws2.cell(row=1, column=ci, value=h).font = Font(bold=True)
            for ri, row in enumerate(data_rows, 2):
                for ci, h in enumerate(headers, 1):
                    val = row.get(h, '')
                    ws2.cell(row=ri, column=ci, value=val if val is not None else '')
            wb.save(filepath)

        return send_file(filepath, as_attachment=True, download_name=filename)

    except Exception as e:
        app.logger.error(f"[EXPORT ERROR] {e}")
        return jsonify({'success': False, 'error': f'导出失败: {str(e)}'})


@app.route('/api/debug/duckdb-info', methods=['GET'])
def debug_duckdb_info():
    """调试: 查看 DuckDB 表状态"""
    try:
        engine = get_duckdb_engine()
        info = {'session_id': session.get('session_id'), 'duckdb_imported': session.get('duckdb_imported')}
        for tbl in ['data', 'balance_data']:
            if engine.table_exists(tbl):
                cols = engine.get_schema(tbl)
                cnt = engine.get_total_rows(tbl)
                info[tbl] = {'exists': True, 'columns': cols, 'row_count': cnt}
            else:
                info[tbl] = {'exists': False}
        return jsonify({'success': True, 'info': info})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in Config.ALLOWED_EXTENSIONS


# ══════════════════════════════════════════════════════════
#  Dify 客户端工厂函数（硬编码配置，无需用户设置）
# ══════════════════════════════════════════════════════════

def _get_dify_client():
    """获取主 Dify Workflow 客户端（SQL生成、字段映射、报表清洗等）

    配置在 config.py 的 DIFY_MAIN_* 变量中，修改后重启应用生效。
    """
    return DifyClient(Config.DIFY_MAIN_BASE_URL, Config.DIFY_MAIN_API_KEY)


def _get_review_dify_client():
    """获取复核 Dify Workflow 客户端（SQL 代码复核审查）

    配置在 config.py 的 DIFY_REVIEW_* 变量中，修改后重启应用生效。
    """
    return DifyClient(Config.DIFY_REVIEW_BASE_URL, Config.DIFY_REVIEW_API_KEY)


def cleanup_stale_files():
    """启动时清理超过24小时的临时文件"""
    now = time.time()
    cutoff = 24 * 3600
    deleted = []

    # 清理 temp/db/*.db 和 *.wal
    db_dir = app.config['DUCKDB_DIR']
    if os.path.isdir(db_dir):
        for fname in os.listdir(db_dir):
            fpath = os.path.join(db_dir, fname)
            if not os.path.isfile(fpath):
                continue
            if fname.endswith('.db') or fname.endswith('.wal'):
                if now - os.path.getmtime(fpath) > cutoff:
                    try:
                        os.remove(fpath)
                        deleted.append(fpath)
                    except Exception as e:
                        app.logger.error(f"[CLEANUP STARTUP] 删除失败: {fpath}, {e}")

    # 清理 temp/result_*.json 和 temp/export_*.*
    upload_dir = app.config['UPLOAD_FOLDER']
    if os.path.isdir(upload_dir):
        for fname in os.listdir(upload_dir):
            fpath = os.path.join(upload_dir, fname)
            if not os.path.isfile(fpath):
                continue
            if not (fname.startswith('result_') or fname.startswith('export_')):
                continue
            if now - os.path.getmtime(fpath) > cutoff:
                try:
                    os.remove(fpath)
                    deleted.append(fpath)
                except Exception as e:
                    app.logger.error(f"[CLEANUP STARTUP] 删除失败: {fpath}, {e}")

    if deleted:
        app.logger.info(f"[CLEANUP STARTUP] 共清理 {len(deleted)} 个过期文件")
        for p in deleted:
            app.logger.info(f"[CLEANUP STARTUP]   已删除: {p}")
    else:
        app.logger.info("[CLEANUP STARTUP] 无需清理（无过期文件）")


if __name__ == '__main__':
    cleanup_stale_files()
    app.run(debug=True, port=5003)