"""
财务报表清洗模块 — 资产负债表/利润表导入、AI 结构检测、规则提取

用法:
    cleaner = ReportCleaner()

    # 1. 加载文件
    sheets = cleaner.load_file("path/to/file.xlsx")

    # 2. AI 检测（读前 10 行）
    meta = cleaner.ai_detect(sheet_name, api_key, provider="deepseek")

    # 3. 规则提取
    result = cleaner.extract_by_meta(sheet_name, meta)
"""

from __future__ import annotations

import os
import re
import json as json_module
import tempfile
import requests


class ReportCleaner:
    """财务报表清洗器"""

    # 需要跳过的行关键词
    DEFAULT_SKIP_KEYWORDS = ["合计", "总计", "小计", "单位负责人", "制表人",
                             "财务负责人", "会计主管", "注：", "注:", "备注"]

    # 空值/零值关键词
    NULL_VALUES = {"", "nan", "none", "null", "n/a", "#n/a", "\\n"}

    def __init__(self):
        self._filepath = None
        self._sheets = {}            # {name: {rows: [[...]], total_rows, total_cols}}
        self._raw_book = None        # openpyxl workbook (for data_only reading)

    # ── 文件加载 ──────────────────────────────────────────

    def load_file(self, filepath: str) -> list:
        """
        加载 Excel 文件，返回 sheet 列表。
        支持：.xlsx (openpyxl) → .xls (xlrd) → 后缀名错的 xlsx 重试
        """
        self._filepath = filepath
        ext = os.path.splitext(filepath)[1].lower()

        # 先试 openpyxl
        if ext == '.xlsx':
            ok = self._load_with_openpyxl(filepath)
            if ok:
                return self._sheet_list()
            raise ValueError(f"无法用 openpyxl 读取: {filepath}")

        # 非 .xlsx → 先试 openpyxl（覆盖后缀名错的 xls），不行再试 xlrd
        ok = self._load_with_openpyxl(filepath)
        if ok:
            return self._sheet_list()

        ok = self._load_with_xlrd(filepath)
        if ok:
            return self._sheet_list()

        # 最后招：复制一份改后缀名再试 openpyxl
        try:
            tmp = filepath + ".report_cleaner_tmp.xlsx"
            import shutil
            shutil.copy2(filepath, tmp)
            ok = self._load_with_openpyxl(tmp)
            os.unlink(tmp)
            if ok:
                return self._sheet_list()
        except Exception:
            pass

        raise ValueError(f"无法读取文件（不支持格式或文件已损坏）: {filepath}")

    def _load_with_openpyxl(self, filepath: str) -> bool:
        """用 openpyxl 加载"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True)
            wb_data = load_workbook(filepath, data_only=True, read_only=True)
            self._raw_book = wb_data
        except Exception:
            return False

        self._sheets = {}
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            max_col = 0
            for row in ws.iter_rows(values_only=True):
                r = [str(c) if c is not None else '' for c in row]
                max_col = max(max_col, len(r))
                rows.append(r)
            self._sheets[name] = {
                'rows': rows,
                'total_rows': len(rows),
                'total_cols': max_col,
            }
            # 用 data_only 模式取公式计算值
            if wb_data:
                try:
                    d_ws = wb_data[name]
                    d_rows = []
                    for row in d_ws.iter_rows(values_only=True):
                        d_rows.append([str(c) if c is not None else '' for c in row])
                    # 替换 formula 行的值
                    for ri in range(min(len(d_rows), len(rows))):
                        for ci in range(min(len(d_rows[ri]), len(rows[ri]))):
                            if d_rows[ri][ci] and d_rows[ri][ci] != 'None':
                                self._sheets[name]['rows'][ri][ci] = d_rows[ri][ci]
                except Exception:
                    pass

        wb.close()
        if wb_data:
            wb_data.close()
        return True

    def _load_with_xlrd(self, filepath: str) -> bool:
        """用 xlrd 加载 .xls"""
        try:
            import xlrd
            wb = xlrd.open_workbook(filepath)
        except Exception:
            return False

        self._sheets = {}
        for name in wb.sheet_names():
            ws = wb.sheet_by_name(name)
            rows = []
            for i in range(ws.nrows):
                row = []
                for j in range(ws.ncols):
                    val = ws.cell_value(i, j)
                    if ws.cell_type(i, j) == 0:  # empty
                        row.append('')
                    else:
                        row.append(str(val) if val is not None else '')
                rows.append(row)
            self._sheets[name] = {
                'rows': rows,
                'total_rows': ws.nrows,
                'total_cols': ws.ncols,
            }
        return True

    def _sheet_list(self) -> list:
        return [{'name': n, 'total_rows': s['total_rows'], 'total_cols': s['total_cols']}
                for n, s in self._sheets.items()]

    def get_sheet_names(self) -> list:
        return list(self._sheets.keys())

    def get_total_rows(self, sheet_name: str) -> int:
        return self._sheets.get(sheet_name, {}).get('total_rows', 0)

    # ── 预览 ──────────────────────────────────────────────

    def preview_raw(self, sheet_name: str, nrows: int = 10) -> list:
        """取前 N 行原始数据"""
        rows = self._sheets.get(sheet_name, {}).get('rows', [])
        return rows[:nrows]

    def preview_all(self, sheet_name: str) -> list:
        """取全部行"""
        return self._sheets.get(sheet_name, {}).get('rows', [])

    # ── AI 结构检测 ───────────────────────────────────────

    def ai_detect(self, sheet_name: str, api_key: str,
                  provider: str = "deepseek",
                  api_url: str = None,
                  model: str = None) -> dict:
        """
        调用 AI 检测报表结构。
        返回结构化 meta dict。
        """
        from config import Config
        cfg = Config.AI_PROVIDERS.get(provider, Config.AI_PROVIDERS["deepseek"])
        url = api_url or cfg.get("api_url", Config.DEEPSEEK_API_URL)
        model_name = model or cfg.get("model", Config.DEEPSEEK_MODEL)

        preview_rows = self.preview_raw(sheet_name, nrows=10)
        rows_text = "\n".join(
            f"第{r+1}行: " + " | ".join(
                f'"{c}"' if c.strip() else "(空)"
                for c in row
            )
            for r, row in enumerate(preview_rows)
        )

        prompt = f"""你是一位财务报表格式分析专家。请分析以下原始数据，识别其结构和格式。

## 工作表名称: {sheet_name}
## 前10行原始数据（每行用 | 分隔单元格）：
{rows_text}

## 要求
请分析并返回 JSON（不要 ```json 包裹，只返回纯 JSON）：

{{
  "report_type": "balance_sheet" | "income_statement" | "other",
  "confidence": 0.95,
  "header_rows_count": 3,
  "data_start_row": 4,
  "data_end_row": 58,
  "layout_type": "left_right_split" | "single_side",
  "columns": [
    {{"index": 0, "name": "资  产", "standard_field": "project_name", "side": "left"}},
    {{"index": 1, "name": "行次", "standard_field": "line_no", "side": "left"}},
    {{"index": 2, "name": "期末余额", "standard_field": "period_end", "side": "left"}},
    {{"index": 3, "name": "年初余额", "standard_field": "period_begin", "side": "left"}},
    {{"index": 4, "name": "负债和所有者权益", "standard_field": "project_name", "side": "right"}},
    {{"index": 5, "name": "行次", "standard_field": "line_no", "side": "right"}},
    {{"index": 6, "name": "期末余额", "standard_field": "period_end", "side": "right"}},
    {{"index": 7, "name": "年初余额", "standard_field": "period_begin", "side": "right"}}
  ],
  "skip_keywords": ["合计", "总计", "单位负责人", "制表人", "注：", "注:", "财务负责人"],
  "report_period": "2025年12月",
  "company_name": "长春英纳法富晟汽车天窗系统有限公司"
}}

## 判断规则
1. **资产负债表**: 标题含"资产负债表"，左右分栏常见（左资产右权益），含流动资产/非流动资产/流动负债等大类
2. **利润表**: 标题含"利润表"，含营业收入/营业成本/净利润等行项目
3. **layout_type**: 如果第3行之后中间列又出现项目名称类文本且总列数>=6则 left_right_split，否则 single_side
4. **standard_field 可选值**: project_name（项目名称）、line_no（行次）、period_end（期末余额）、period_begin（年初余额/期初余额）、month_amount（本月数）、year_amount（本年累计数）、last_year_amount（上年同期）、ignore（忽略列）
5. **side**: left 或 right，single_side 时全填 "center"
6. **data_end_row**: 数据实际结束的行号（在签名行/注释行之前）
7. **skip_keywords**: 包含此行关键字的行应被过滤
8. **confidence**: 0-1，对自己判断的确认程度
"""

        # 调用 AI API
        payload = {
            "model": model_name,
            "messages": [
                {"role": "system", "content": "你是一个专业的财务报表格式分析专家，返回严格 JSON 格式的分析结果。"},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.1,
            "max_tokens": 2000,
        }

        try:
            resp = requests.post(
                url,
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                },
                json=payload,
                timeout=30,
            )
            if resp.status_code != 200:
                err = resp.json().get('error', {}).get('message', str(resp.status_code))
                return {"success": False, "error": f"API 返回错误: {err}"}

            content = resp.json()['choices'][0]['message']['content']
        except requests.exceptions.Timeout:
            return {"success": False, "error": "API 请求超时，请重试"}
        except requests.exceptions.ConnectionError:
            return {"success": False, "error": "网络连接失败，请检查网络"}
        except Exception as e:
            return {"success": False, "error": f"API 调用异常: {e}"}

        # 解析 JSON
        meta = self._parse_json_response(content)
        if meta is None:
            return {"success": False, "error": "AI 返回格式异常，无法解析"}

        meta["success"] = True
        return meta

    def _parse_json_response(self, text: str) -> dict | None:
        """从 AI 响应中提取 JSON"""
        text = text.strip()
        # 尝试直接解析
        try:
            return json_module.loads(text)
        except json_module.JSONDecodeError:
            pass

        # 尝试匹配 ```json ... ```
        m = re.search(r'```(?:json)?\s*\n?(.*?)\n?```', text, re.DOTALL)
        if m:
            try:
                return json_module.loads(m.group(1))
            except json_module.JSONDecodeError:
                pass

        # 尝试匹配第一个 { 到最后一个 }
        m = re.search(r'\{.*\}', text, re.DOTALL)
        if m:
            try:
                return json_module.loads(m.group(0))
            except json_module.JSONDecodeError:
                pass

        return None

    # ── 规则提取 ──────────────────────────────────────────

    def extract_by_meta(self, sheet_name: str, meta: dict) -> dict:
        """
        根据 AI 检测结果，规则提取清洗后数据。
        返回: {success, columns, data, row_count, removed, warnings}
        """
        all_rows = self._sheets.get(sheet_name, {}).get('rows', [])
        if not all_rows:
            return {"success": False, "error": "Sheet 无数据"}

        try:
            data_start = meta.get("data_start_row", 4) - 1  # 转 0-indexed
            data_end = meta.get("data_end_row", len(all_rows))  # 已经是 1-indexed
            columns_def = meta.get("columns", [])
            layout = meta.get("layout_type", "single_side")
            skip_kw = meta.get("skip_keywords", self.DEFAULT_SKIP_KEYWORDS)
        except Exception as e:
            return {"success": False, "error": f"检测结果 meta 解析失败: {e}"}

        if not columns_def:
            return {"success": False, "error": "未检测到列定义"}

        # 收集需要过滤的行关键词
        skip_set = set(kw.strip() for kw in skip_kw if kw.strip())
        # 合并默认关键词
        skip_set.update(self.DEFAULT_SKIP_KEYWORDS)
        # 也过滤纯数字行次的行（如只有行号没有项目名）
        section_markers = set()  # "流动资产：", "流动负债：" 等大类标题

        # 数据行
        data_rows = all_rows[data_start:data_end]

        # 分组列：left / right / center
        left_cols = [c for c in columns_def if c.get("side") == "left"]
        right_cols = [c for c in columns_def if c.get("side") == "right"]
        center_cols = [c for c in columns_def if c.get("side") in ("center", "")]

        # 标准化列名映射
        std_fields_map = {}  # {col_index: standard_field_name}
        field_order = []     # 按 index 排序的标准字段名

        # 先处理 center 或 left
        cols_to_use = center_cols if center_cols else left_cols
        for c in cols_to_use:
            idx = c.get("index")
            sf = c.get("standard_field", "")
            if sf != "ignore":
                std_fields_map[idx] = sf
                if sf not in field_order:
                    field_order.append(sf)

        # 确定左右分栏的分隔列
        split_col = None
        if layout == "left_right_split" and right_cols:
            split_col = min(c.get("index") for c in right_cols)

        # 提取
        cleaned = []
        removed = {"summary": 0, "signature": 0, "annotation": 0, "empty": 0, "section_header": 0}

        for row in data_rows:
            # 检查是否空行
            if self._is_empty_row(row):
                removed["empty"] += 1
                continue

            # 检查是否注释行
            row_text = " ".join(str(c) for c in row if str(c).strip())
            if any(row_text.startswith(kw) for kw in ["注：", "注:", "备注"]):
                removed["annotation"] += 1
                continue
            # 签名行
            if any(kw in row_text for kw in ["单位负责人", "制表人", "会计主管", "财务负责人"]):
                removed["signature"] += 1
                continue
            # 汇总行
            if any(kw in row_text for kw in ["合计", "总计"]):
                removed["summary"] += 1
                continue
            # 小计行
            row_stripped = row[0].strip() if row and len(row) > 0 else ''
            if row_stripped in ("小计", "小  计") or row_stripped.endswith("小计"):
                removed["summary"] += 1
                continue

            # 大类标题行（"流动资产：" 等）— 保留但标记
            is_section_header = False
            if row_stripped and any(
                row_stripped.startswith(prefix)
                for prefix in ["流动资产", "非流动资产", "流动负债", "非流动负债",
                               "所有者权益", "资产", "负债", "权益"]
            ) and row_stripped.endswith("："):
                is_section_header = True

            if is_section_header:
                # 作为 section 行输出
                record = {"项目名称": row_stripped.rstrip("："), "报表侧": "left" if layout == "left_right_split" else "center"}
                for sf in field_order:
                    if sf not in record:
                        record[sf] = ""
                cleaned.append(record)
                continue

            # 普通数据行 — 按列定义提取
            if layout == "left_right_split" and split_col:
                # 左右分栏：左右各生成一行
                left_row = self._extract_row_values(row, left_cols, std_fields_map)
                right_row = self._extract_row_values(row, right_cols, std_fields_map)

                if left_row and self._has_item_name(left_row):
                    left_row["报表侧"] = "资产"
                    if "line_no" not in left_row or not left_row["line_no"]:
                        left_row["line_no"] = ""
                    cleaned.append(left_row)

                if right_row and self._has_item_name(right_row):
                    right_row["报表侧"] = "负债和所有者权益"
                    if "line_no" not in right_row or not right_row["line_no"]:
                        right_row["line_no"] = ""
                    cleaned.append(right_row)
            else:
                record = self._extract_row_values(row, cols_to_use, std_fields_map)
                if record and self._has_item_name(record):
                    record["报表侧"] = ""
                    cleaned.append(record)

        # 数值清洗
        amount_fields = ["period_end", "period_begin", "month_amount",
                         "year_amount", "last_year_amount"]
        for rec in cleaned:
            for af in amount_fields:
                if af in rec and rec[af]:
                    rec[af] = self._clean_number(rec[af])

        # 列名映射为中文
        FIELD_NAMES = {
            "project_name": "项目名称",
            "line_no": "行次",
            "period_end": "期末余额",
            "period_begin": "年初余额",
            "month_amount": "本月金额",
            "year_amount": "本年累计金额",
            "last_year_amount": "上年同期累计数",
            "side": "报表侧",
        }

        # 翻译字段名
        translated = []
        for rec in cleaned:
            t = {}
            for k, v in rec.items():
                cn_key = FIELD_NAMES.get(k, k)
                t[cn_key] = v
            translated.append(t)

        # 检查是否提取出数据
        if not translated:
            return {"success": False, "error": "未提取到有效数据行，请检查 AI 检测结果是否正确"}

        # 根据报表类型只保留需要的列
        report_type = meta.get("report_type", "")
        if report_type == "income_statement":
            # 利润表：很多 ERP 把金额列标为"本年累计金额"，统一输出为"期末余额"
            KEEP_COLUMNS_BS = ["项目名称", "期末余额"]
            # 检查数据中是否有"期末余额"，如果没有，从其它金额列取第一个补齐
            has_period_end = any("期末余额" in rec for rec in translated)
            if not has_period_end:
                # 找到第一个存在的金额列
                for alt in ["本年累计金额", "本月金额", "上年同期累计数"]:
                    if any(alt in rec for rec in translated):
                        for rec in translated:
                            if alt in rec:
                                rec["期末余额"] = rec.pop(alt)
                        break
            want_cols = KEEP_COLUMNS_BS
        elif report_type == "balance_sheet":
            want_cols = ["项目名称", "年初余额", "期末余额"]
        else:
            want_cols = None

        if want_cols:
            filtered = []
            for rec in translated:
                f = {}
                for col in want_cols:
                    f[col] = rec.get(col, "")
                filtered.append(f)
        else:
            filtered = translated

        return {
            "success": True,
            "columns": list(filtered[0].keys()),
            "data": filtered,
            "row_count": len(translated),
            "removed": removed,
            "report_type": meta.get("report_type", "unknown"),
            "report_period": meta.get("report_period", ""),
            "company_name": meta.get("company_name", ""),
        }

    def _extract_row_values(self, row: list, col_defs: list, fields_map: dict) -> dict:
        """从行中按列定义提取字段值"""
        record = {}
        has_value = False
        for cd in col_defs:
            idx = cd.get("index")
            sf = cd.get("standard_field", "")
            if sf == "ignore":
                continue
            if idx < len(row):
                val = row[idx].strip()
            else:
                val = ""
            record[sf] = val
            if sf == "project_name" and val:
                has_value = True
        if not has_value:
            return None
        return record

    def _has_item_name(self, record: dict) -> bool:
        """检查是否有项目名称"""
        name = record.get("project_name", "")
        return bool(name.strip()) if name else False

    def _is_empty_row(self, row: list) -> bool:
        """判断是否空行"""
        return all(str(c).strip() in self.NULL_VALUES for c in row)

    def _clean_number(self, val: str):
        """清洗数值：去逗号、去空格、空转空"""
        val = str(val).strip()
        if not val or val.lower() in ("nan", "none", "null", "n/a", "#n/a"):
            return ""
        val = val.replace(",", "").replace("，", "").replace(" ", "")
        try:
            n = float(val)
            if n == int(n):
                return int(n)
            return n
        except ValueError:
            return val  # 无法转数字的保留原文

    # ── 导出 ──────────────────────────────────────────────

    def export_to_excel(self, sheet_name: str, meta: dict,
                        filepath: str = None) -> str:
        """
        提取并导出清洗后数据到 Excel。
        不传 filepath 则返回临时文件路径。
        """
        result = self.extract_by_meta(sheet_name, meta)
        if not result.get("success"):
            raise ValueError(result.get("error", "提取失败"))

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "清洗后数据"

        # 表头样式
        hfont = Font(bold=True, size=11, color="FFFFFF")
        hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        h_align = Alignment(horizontal="center", vertical="center")

        columns = result["columns"]
        data = result["data"]

        # 写入公司名/期间信息
        row_idx = 1
        if meta.get("company_name"):
            ws.cell(row=row_idx, column=1, value="公司名称").font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=meta["company_name"])
            row_idx += 1
        if meta.get("report_period"):
            ws.cell(row=row_idx, column=1, value="报表期间").font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=meta["report_period"])
            row_idx += 1
        ws.cell(row=row_idx, column=1, value=f"报表类型: {meta.get('report_type', '')}").font = Font(bold=True)
        row_idx += 2

        # 表头
        for ci, h in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=ci, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = h_align
        row_idx += 1

        # 数据
        amount_fields_set = {"期末余额", "年初余额", "本月金额", "本年累计金额", "上年同期累计数"}
        for rec in data:
            for ci, h in enumerate(columns, 1):
                val = rec.get(h, "")
                cell = ws.cell(row=row_idx, column=ci)

                if h in amount_fields_set and val != "":
                    try:
                        cell.value = float(val)
                        cell.number_format = '#,##0.00'
                    except (ValueError, TypeError):
                        cell.value = val
                else:
                    cell.value = val
            row_idx += 1

        # 列宽
        col_widths = {
            "项目名称": 32, "行次": 8, "期末余额": 18, "年初余额": 18,
            "本月金额": 18, "本年累计金额": 18, "上年同期累计数": 18, "报表侧": 20,
        }
        for ci, h in enumerate(columns, 1):
            ws.column_dimensions[chr(64 + ci)].width = col_widths.get(h, 14)

        if filepath:
            wb.save(filepath)
            return filepath

        fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="report_clean_")
        os.close(fd)
        wb.save(path)
        return path
