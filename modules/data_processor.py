import pandas as pd
import numpy as np
import os
from typing import Dict, List, Any, Optional

class DataProcessor:
    """简化的数据处理模块"""

    def __init__(self, filepath: str, field_mapping: Optional[Dict[str, str]] = None):
        self.filepath = filepath
        self.field_mapping = field_mapping
        self.df = None
        self.file_type = self._detect_file_type()

    def _detect_file_type(self) -> str:
        ext = os.path.splitext(self.filepath)[1].lower()
        if ext in ['.xlsx', '.xls']:
            return 'excel'
        elif ext == '.csv':
            return 'csv'
        else:
            raise ValueError(f"不支持的文件类型: {ext}")

    @staticmethod
    def get_xlsx_sheet_names(filepath: str) -> list:
        """获取 xlsx 文件的所有 sheet 名称"""
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets

    @staticmethod
    def preview_raw(filepath: str, nrows: int = 15) -> dict:
        """
        预览原始数据（不解析为 DataFrame），返回多 sheet 的原始行
        返回值: { sheet_name: { rows: [[...], ...], total_rows: N, total_cols: N } }
        """
        ext = os.path.splitext(filepath)[1].lower()
        if ext in ['.xlsx', '.xls']:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True)
            result = {}
            for name in wb.sheetnames:
                ws = wb[name]
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= nrows:
                        break
                    rows.append([str(c) if c is not None else '' for c in row])
                result[name] = {
                    'rows': rows,
                    'total_rows': ws.max_row,
                    'total_cols': ws.max_column,
                }
            wb.close()
            return result
        else:
            # CSV — 当作单个 sheet 处理
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                rows = []
                for i, row in enumerate(reader):
                    if i >= nrows:
                        break
                    rows.append(row)
            max_cols = max(len(r) for r in rows) if rows else 0
            return {'__csv__': {
                'rows': rows,
                'total_rows': len(rows),
                'total_cols': max_cols,
            }}

    def load_data(self, nrows: Optional[int] = None,
                  sheet_name: Optional[str] = None,
                  header_row: Optional[int] = None) -> pd.DataFrame:
        """加载数据"""
        try:
            if self.file_type == 'excel':
                kwargs = {'header': header_row} if header_row is not None else {}
                if sheet_name:
                    kwargs['sheet_name'] = sheet_name
                self.df = pd.read_excel(self.filepath, nrows=nrows, **kwargs)
            else:
                kwargs = {'header': header_row} if header_row is not None else {}
                # 尝试常见编码，包括带BOM的UTF-8
                encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'latin1', 'cp1252']
                for encoding in encodings:
                    try:
                        self.df = pd.read_csv(self.filepath, nrows=nrows, encoding=encoding, **kwargs)
                        break
                    except UnicodeDecodeError:
                        continue
                if self.df is None:
                    self.df = pd.read_csv(self.filepath, nrows=nrows, encoding='utf-8', errors='ignore', **kwargs)

            # 清理列名
            if self.df is not None and not self.df.empty:
                self.df.columns = self.df.columns.str.strip()
            return self.df
        except Exception as e:
            raise Exception(f"加载数据失败: {str(e)}")

    def get_total_rows(self, sheet_name: Optional[str] = None, header_row: Optional[int] = None) -> int:
        """高效获取总行数（不加载全部数据）"""
        try:
            if self.file_type == 'excel':
                from openpyxl import load_workbook
                wb = load_workbook(self.filepath, read_only=True)
                if sheet_name and sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.active
                offset = (header_row + 1) if header_row is not None else 1  # 0-indexed → 行数
                total = max(0, ws.max_row - offset)
                wb.close()
                return total
            else:
                # CSV: 尝试常见编码
                encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'latin1', 'cp1252']
                for encoding in encodings:
                    try:
                        df = pd.read_csv(self.filepath, encoding=encoding)
                        return len(df)
                    except (UnicodeDecodeError, ValueError):
                        continue
                df = pd.read_csv(self.filepath, encoding='utf-8', errors='ignore')
                return len(df)
        except Exception:
            return 0

    def analyze_columns(self) -> List[Dict[str, Any]]:
        """分析列"""
        if self.df is None:
            self.load_data(nrows=100)

        columns_info = []
        for col in self.df.columns:
            col_data = self.df[col]
            non_null = col_data.count()
            null_count = len(col_data) - non_null
            unique_count = col_data.nunique()

            # 简单类型检测
            data_type = 'text'
            try:
                pd.to_numeric(col_data, errors='coerce')
                if pd.to_numeric(col_data, errors='coerce').count() > len(col_data) * 0.5:
                    data_type = 'number'
            except:
                pass

            # 日期检测 - 更宽松的关键词匹配
            date_keywords = ['日期', 'date', 'time', '时间', '年', '月', '日', 'period', '会计期间', '记账日期', '交易日期']
            if any(keyword in col.lower() for keyword in date_keywords):
                try:
                    # 使用 format='mixed' 处理混合格式的日期字符串
                    # errors='coerce' 将无法解析的值转为 NaT
                    converted = pd.to_datetime(col_data, format='mixed', errors='coerce')
                    if converted.notna().sum() > 0:
                        data_type = 'date'
                except Exception as e:
                    print(f"日期检测异常 - 列 '{col}', 样本值 '{col_data.iloc[0] if len(col_data) > 0 else None}': {e}")
                    import traceback
                    traceback.print_exc()

            sample = col_data.iloc[0] if len(col_data) > 0 else None
            if pd.isna(sample):
                sample = None
            elif isinstance(sample, (np.integer, np.int64, np.int32, np.int16, np.int8)):
                sample = int(sample)
            elif isinstance(sample, (np.floating, np.float64, np.float32, np.float16)):
                sample = float(sample)
            elif isinstance(sample, (pd.Timestamp, np.datetime64)):
                sample = str(sample)[:10]

            # 截断过长的样本值以减少会话大小
            if isinstance(sample, str) and len(sample) > 100:
                sample = sample[:100] + '...'

            columns_info.append({
                'name': col,
                'type': data_type,
                'non_null_count': int(non_null),
                'null_count': int(null_count),
                'unique_count': int(unique_count),
                'sample': sample
            })

        return columns_info

    def get_preview_data(self, rows: int = 5) -> List[Dict[str, Any]]:
        """获取预览数据（默认5行以减少会话大小）"""
        if self.df is None:
            self.load_data(rows)

        preview_df = self.df.head(rows)
        preview_data = []
        for _, row in preview_df.iterrows():
            row_dict = {}
            for col in preview_df.columns:
                val = row[col]
                if pd.isna(val):
                    row_dict[col] = None
                elif isinstance(val, (pd.Timestamp, np.datetime64)):
                    row_dict[col] = str(val)[:10]
                elif isinstance(val, (int, np.integer, float, np.float64)):
                    row_dict[col] = float(val) if isinstance(val, float) else int(val)
                else:
                    row_dict[col] = str(val)
            preview_data.append(row_dict)

        return preview_data

    def process(self, field_mapping: Optional[Dict[str, str]] = None,
                sheet_name: Optional[str] = None,
                header_row: Optional[int] = None) -> Dict[str, Any]:
        """处理数据文件"""
        try:
            # 加载数据用于分析
            self.load_data(nrows=100, sheet_name=sheet_name, header_row=header_row)

            columns_info = self.analyze_columns()
            preview_data = self.get_preview_data()
            total_rows = self.get_total_rows(sheet_name=sheet_name, header_row=header_row)

            result = {
                'success': True,
                'filename': os.path.basename(self.filepath),
                'row_count': int(total_rows),
                'column_count': int(len(self.df.columns)),
                'fields': columns_info,
                'preview': preview_data,
                'has_mapping': field_mapping is not None,
                'field_mapping': field_mapping,
                'mapped_fields': self._apply_field_mapping_to_fields(columns_info, field_mapping),
                'mapped_preview': self._apply_field_mapping_to_preview(preview_data, field_mapping)
            }

            return result
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'filename': os.path.basename(self.filepath)
            }

    def clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """清理数据 - 修复数值/NaN/类型问题，确保沙箱执行稳定"""
        cleaned_df = df.copy()

        for col in cleaned_df.columns:
            # 对 object 类型列：用空字符串填充 NaN（避免 str 操作失败）
            if cleaned_df[col].dtype == 'object':
                try:
                    # 先尝试数值转换（对纯数字列有效）
                    numeric_col = pd.to_numeric(cleaned_df[col], errors='coerce')
                    # 如果大部分值能转换为数字，则保留为数值类型（NaN 填充 0）
                    if numeric_col.notna().sum() > len(cleaned_df) * 0.5:
                        cleaned_df[col] = numeric_col.fillna(0)
                    else:
                        # 否则作为字符串处理，NaN 填充为空字符串
                        cleaned_df[col] = cleaned_df[col].fillna('').astype(str).str.strip()
                except Exception:
                    cleaned_df[col] = cleaned_df[col].fillna('').astype(str).str.strip()

            # 对数值列：NaN 填充为 0
            elif pd.api.types.is_bool_dtype(cleaned_df[col]):
                # 布尔列先转字符串（必须放在 is_numeric_dtype 之前，否则 bool 会被当作数值处理）
                cleaned_df[col] = cleaned_df[col].astype(str)
            elif pd.api.types.is_numeric_dtype(cleaned_df[col]):
                cleaned_df[col] = cleaned_df[col].fillna(0)

        return cleaned_df

    def import_full_data(self, duckdb_engine, table_name: str = 'data',
                         sheet_name: Optional[str] = None,
                         header_row: Optional[int] = None) -> int:
        """
        将完整数据导入 DuckDB

        Args:
            duckdb_engine: DuckDBEngine 实例
            table_name: 目标表名
            sheet_name: Excel sheet 名称
            header_row: 表头行号（0-indexed）

        Returns:
            导入的行数
        """
        rename_mapping = None
        if self.field_mapping:
            rename_mapping = {v: k for k, v in self.field_mapping.items()}

        if self.file_type == 'excel':
            return duckdb_engine.import_xlsx(
                self.filepath, table_name, rename_mapping,
                sheet_name=sheet_name, header_row=header_row,
            )
        else:
            return duckdb_engine.import_csv(
                self.filepath, table_name, rename_mapping=rename_mapping,
                header_row=header_row,
            )

    def _apply_field_mapping_to_fields(self, fields_info, field_mapping):
        """应用字段映射到字段信息"""
        if not field_mapping:
            return fields_info

        # field_mapping 格式为 {标准字段名: 源字段名}，反转后得到 {源字段名: 标准字段名}
        reverse_mapping = {v: k for k, v in field_mapping.items()}

        mapped_fields = []
        for field in fields_info:
            field_name = field.get('name', '')
            if field_name in reverse_mapping:
                # 创建字段副本并更新为标准名称
                mapped_field = field.copy()
                mapped_field['name'] = reverse_mapping[field_name]
                mapped_field['original_name'] = field_name  # 保留原始名称
                mapped_fields.append(mapped_field)
            else:
                # 字段没有映射，保持原样
                mapped_fields.append(field.copy())

        return mapped_fields

    def _apply_field_mapping_to_preview(self, preview_data, field_mapping):
        """应用字段映射到预览数据"""
        if not field_mapping or not preview_data:
            return preview_data

        # field_mapping 格式为 {标准字段名: 源字段名}，反转后得到 {源字段名: 标准字段名}
        reverse_mapping = {v: k for k, v in field_mapping.items()}

        mapped_preview = []
        for row in preview_data:
            mapped_row = {}
            for key, value in row.items():
                if key in reverse_mapping:
                    mapped_key = reverse_mapping[key]
                else:
                    mapped_key = key
                mapped_row[mapped_key] = value
            mapped_preview.append(mapped_row)

        return mapped_preview