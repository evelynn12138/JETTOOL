"""
会计科目向量数据库 — 语义查询工具

用法:
    python query_db.py "固定资产的核算方法"
    python query_db.py "哪些科目会影响利润表" --top 10
    python query_db.py --interactive
    python query_db.py --code 1001              # 按编号精确查找
    python query_db.py --name "短期"             # 按名称模糊搜索
    python query_db.py --export subjects.xlsx    # 导出全部数据到 Excel
"""

import argparse
import os
import sys

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_DIR = os.path.join(BASE_DIR, "chroma_db")

sys.path.insert(0, BASE_DIR)
from data import ACCOUNTING_SUBJECTS, search_by_code, search_by_name


def get_collection():
    """加载已构建的 ChromaDB 集合"""
    import chromadb
    from chromadb import EmbeddingFunction, Documents, Embeddings
    from sentence_transformers import SentenceTransformer

    if not os.path.exists(DB_DIR):
        print("错误: 数据库不存在。请先运行 python build_db.py")
        sys.exit(1)

    class ChineseEmbeddingFunction(EmbeddingFunction):
        def __init__(self):
            model_name = "shibing624/text2vec-base-chinese"
            print(f"[加载模型] {model_name}")
            self.model = SentenceTransformer(model_name, trust_remote_code=True)

        def __call__(self, input: Documents) -> Embeddings:
            return self.model.encode(
                list(input), show_progress_bar=False, normalize_embeddings=True
            ).tolist()

    client = chromadb.PersistentClient(path=DB_DIR)
    return client.get_collection("accounting_subjects", embedding_function=ChineseEmbeddingFunction())


def semantic_search(collection, query: str, top_k: int = 5):
    """语义搜索"""
    results = collection.query(
        query_texts=[query],
        n_results=min(top_k, collection.count()),
    )

    rows = []
    for i, (doc, meta, dist) in enumerate(
        zip(results["documents"][0], results["metadatas"][0], results["distances"][0])
    ):
        score = 1 - dist  # cosine similarity
        rows.append({
            "rank": i + 1,
            "code": meta["code"],
            "name": meta["name"],
            "category": meta["category"],
            "report_item": meta["report_item"],
            "similarity": f"{score:.4f}",
        })

    return rows


def print_table(rows):
    """打印格式化表格"""
    if not rows:
        print("(无结果)")
        return

    # 列宽
    name_w = max(len(r.get("name", "")) for r in rows)
    code_w = max(len(r.get("code", "")) for r in rows)
    cat_w = max(len(r.get("category", "")) for r in rows)
    report_w = max(len(r.get("report_item", "")) for r in rows)
    score_w = 8

    name_w = max(name_w, 8)
    code_w = max(code_w, 8)
    cat_w = max(cat_w, 6)
    report_w = max(report_w, 10)

    sep = "-" * (code_w + name_w + cat_w + report_w + score_w + 20)

    print(f"  {'排名':<4} {'编号':<{code_w}} {'科目名称':<{name_w}} {'类别':<{cat_w}} {'报表项目':<{report_w}} {'相似度':<{score_w}}")
    print(sep)
    for r in rows:
        print(f"  {r.get('rank', ''):<4} {r.get('code', ''):<{code_w}} "
              f"{r.get('name', ''):<{name_w}} {r.get('category', ''):<{cat_w}} "
              f"{r.get('report_item', ''):<{report_w}} {r.get('similarity', ''):<{score_w}}")

    if len(rows) > 0 and "usage" in rows[0]:
        print()
        for r in rows:
            print(f"  ── {r['code']} {r['name']} ──")
            print(f"     使用说明: {r['usage']}")
            print()


def interactive_mode(collection):
    """交互式查询"""
    print("\n=== 会计科目向量数据库 — 交互查询模式 ===")
    print("输入查询内容（自然语言），输入 /q 退出\n")

    while True:
        try:
            q = input("🔍 ").strip()
        except (EOFError, KeyboardInterrupt):
            break

        if not q:
            continue
        if q in ("/q", "/quit", "/exit"):
            break
        if q.startswith("/"):
            cmd, *args = q[1:].split(maxsplit=1)
            if cmd == "code" and args:
                rows = search_by_code(args[0])
            elif cmd == "name" and args:
                rows = search_by_name(args[0])
            elif cmd == "top" and args:
                n = int(args[0])
                rows = semantic_search(collection, "会计科目", top_k=n)
            elif cmd == "help":
                print("  /code 编号  — 按科目编号精确查找")
                print("  /name 名称  — 按科目名称模糊搜索")
                print("  /top N      — 列出前 N 条科目")
                print("  /help       — 显示帮助")
                print("  /q          — 退出")
                continue
            else:
                print(f"未知命令: {cmd}")
                continue
            print_table(rows)
            continue

        # 自然语言语义搜索
        if len(q) > 200:
            print("查询过长，请精简")
            continue

        rows = semantic_search(collection, q, top_k=5)
        # 补充 usage 字段
        full_rows = []
        for r in rows:
            subjects = search_by_code(r["code"])
            if subjects:
                r["usage"] = subjects[0]["usage"]
            full_rows.append(r)
        print_table(full_rows)


def export_to_excel(filepath: str):
    """导出全部科目到 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "会计科目全表"

    headers = ["顺序号", "编号", "科目名称", "类别", "使用说明", "对应报表项目"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for ri, s in enumerate(ACCOUNTING_SUBJECTS, 2):
        ws.cell(row=ri, column=1, value=s["order"])
        ws.cell(row=ri, column=2, value=s["code"])
        ws.cell(row=ri, column=3, value=s["name"])
        ws.cell(row=ri, column=4, value=s["category"])
        ws.cell(row=ri, column=5, value=s["usage"])
        ws.cell(row=ri, column=6, value=s["report_item"])

    # 列宽
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 30

    wb.save(filepath)
    print(f"已导出 {len(ACCOUNTING_SUBJECTS)} 条记录到: {filepath}")


def main():
    parser = argparse.ArgumentParser(description="会计科目向量数据库查询工具")
    parser.add_argument("query", nargs="?", default=None, help="自然语言查询")
    parser.add_argument("--top", type=int, default=5, help="返回 top-K 结果（默认 5）")
    parser.add_argument("--code", type=str, default=None, help="按科目编号精确查找")
    parser.add_argument("--name", type=str, default=None, help="按科目名称模糊搜索")
    parser.add_argument("--interactive", "-i", action="store_true", help="交互模式")
    parser.add_argument("--export", type=str, default=None,
                        help="导出全部数据到 Excel 文件路径")
    args = parser.parse_args()

    # 优先级: export > code > name > query > interactive

    if args.export:
        export_to_excel(args.export)
        return

    # 按编号查找（不需要向量 DB）
    if args.code:
        rows = search_by_code(args.code)
        subjects = [{"code": s["code"], "name": s["name"], "category": s["category"],
                      "report_item": s["report_item"]} for s in rows]
        print_table(subjects)
        if rows:
            print(f"  使用说明: {rows[0]['usage']}")
        return

    # 按名称模糊搜索（不需要向量 DB）
    if args.name:
        rows = search_by_name(args.name)
        subjects = [{"code": s["code"], "name": s["name"], "category": s["category"],
                      "report_item": s["report_item"]} for s in rows]
        print_table(subjects)
        return

    # 需要向量 DB
    collection = get_collection()

    if args.query:
        rows = semantic_search(collection, args.query, top_k=args.top)
        full_rows = []
        for r in rows:
            subj = search_by_code(r["code"])
            if subj:
                r["usage"] = subj[0]["usage"]
            full_rows.append(r)
        print_table(full_rows)
        return

    if args.interactive:
        interactive_mode(collection)
        return

    # 默认无参数 → 交互模式
    interactive_mode(collection)


if __name__ == "__main__":
    main()
